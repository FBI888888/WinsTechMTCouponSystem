"""
聚合bot插件 - 关键词回复、用户管理、日志、授权管理等
"""
import base64
import hmac
import sys
import os
from datetime import datetime, timedelta
import re
import threading
import json
from typing import Dict, Optional, List, Union
import time
import hashlib
from urllib.parse import urlparse, parse_qsl, urlencode
import requests
import warnings

# 禁用SSL警告（用于打包后环境的SSL证书问题）
warnings.filterwarnings('ignore', message='Unverified HTTPS request')
try:
    from urllib3.exceptions import InsecureRequestWarning
    warnings.filterwarnings('ignore', category=InsecureRequestWarning)
except ImportError:
    pass

# 导入数据库相关
try:
    import pymysql
    PYMYSQL_AVAILABLE = True
except ImportError:
    PYMYSQL_AVAILABLE = False

# 导入GenerateCoupon模块
try:
    from GenerateCoupon import gm_main, main as generate_coupon_image, format_voucher_code
    GM_AVAILABLE = True
except ImportError:
    GM_AVAILABLE = False
    gm_main = None
    generate_coupon_image = None
    format_voucher_code = None

# 添加父目录到路径，以便导入plugin_base
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from plugin_base import PluginBase
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, 
    QTextEdit, QLineEdit, QGroupBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QMessageBox, QTabWidget,
    QComboBox, QDateTimeEdit, QSpinBox, QDoubleSpinBox, QMenu, QAction, QInputDialog
)
from PyQt5.QtCore import Qt, QDateTime, QTimer, QMetaObject, Q_ARG
from PyQt5.QtGui import QFont

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ==================== 美团相关辅助函数 ====================

class MeituanRiskControlException(Exception):
    """美团风控异常"""
    def __init__(self, message, general_page_url=None):
        super().__init__(message)
        self.general_page_url = general_page_url


def get_coupon_mt_id(text):
    """从消息文本中提取美团ID"""
    # 使用正则表达式查找字符串
    match = re.search(r"did\%3D(\d+)", text)
    
    # 检查是否找到匹配项
    if match:
        mt_id = match.group(1)
    else:
        match = re.search(r'did=(\d+)', text)
        if match:
            mt_id = match.group(1)
        else:
            mt_id = 0
    return mt_id


def extract_url_params(url):
    """
    从URL中提取指定参数并返回字典
    目标参数: did, ci, uuid, token, prePoiId, preCityId, offset, limit, mypos, dpId, chooseCity, chooseAllCity
    """
    # 解析URL获取查询字符串
    parsed_url = urlparse(url)
    query_params = parse_qsl(parsed_url.query)
    
    # 目标参数列表
    target_keys = {
        'did', 'ci', 'uuid', 'token', 'prePoiId', 'preCityId',
        'offset', 'limit', 'mypos', 'dpId', 'chooseCity', 'chooseAllCity'
    }
    
    # 筛选目标参数
    result_dict = {
        key: value for key, value in query_params
        if key in target_keys
    }
    
    return result_dict


def get_mtgsig_url_get(url1, payload1, server_index=0):
    """
    获取美团签名URL
    
    Args:
        url1: 原始URL
        payload1: 请求参数
        server_index: 服务器索引 (0=首次请求, 1=第一次重试, 2=第二次重试)
    
    Returns:
        str: 签名后的URL
    """
    # 服务器列表：首次请求、第一次重试、第二次重试
    servers = [
        "http://154.8.224.221:3003/get-verify",
        "http://home.jhsrvip.cn:3003/get-verify"
    ]
    
    # 根据server_index选择服务器
    if server_index >= len(servers):
        server_index = len(servers) - 1
    
    url = servers[server_index]
    server_host = url.replace("http://", "").replace("/get-verify", "")
    
    print(f"[get_mtgsig_url_get] 使用服务器[{server_index}]: {server_host}")
    
    max_retries = 2
    retry_count = 0
    
    while retry_count <= max_retries:
        try:
            payload = json.dumps({
                "url": url1,
                "data": payload1
            })
            
            headers = {
                'User-Agent': 'Apifox/1.0.0 (https://apifox.com)',
                'Content-Type': 'application/json',
                'Accept': '*/*',
                'Host': server_host,
                'Connection': 'keep-alive'
            }
            
            response = requests.request(
                "POST", 
                url, 
                headers=headers, 
                data=payload, 
                timeout=15,
                verify=False
            )
            res = response.json()
            return res['url']
            
        except (requests.exceptions.ConnectionError, 
                requests.exceptions.Timeout,
                ConnectionResetError) as e:
            retry_count += 1
            if retry_count > max_retries:
                raise Exception(f"获取签名URL失败，已重试{max_retries}次: {str(e)}")
            time.sleep(1)
            
        except Exception as e:
            retry_count += 1
            if retry_count > max_retries:
                raise Exception(f"获取签名URL异常: {str(e)}")
            time.sleep(1)


def get_mtgsig_w_api(method, mt_url, data, header):
    """获取美团W签名"""
    url = "http://154.8.224.221:3010/get_w_mtgsig"

    payload = json.dumps({
        "method": method,
        "url": mt_url,
        "data": data,
        "header": header
    })
    headers = {
        'User-Agent': 'Apifox/1.0.0 (https://apifox.com)',
        'Content-Type': 'application/json',
        'Accept': '*/*',
        'Host': '154.8.224.221:3010',
        'Connection': 'keep-alive'
    }

    response = requests.request("POST", url, headers=headers, data=payload)
    res = response.json()
    mtgsig = res['data']['mtgsig']
    return mtgsig


def api_get_mt_order_rebate_info(orderViewId, token, userid):
    """查询美团订单返利信息（直接调用get_mt_order_rebate_info）"""
    # 本地API服务器已停用，直接使用美团接口
    return get_mt_order_rebate_info(orderViewId, token, userid)


def get_mt_order_rebate_info(orderViewId, token, userid):
    """查询美团订单返利信息（带重试机制）"""
    url = "https://media.meituan.com/mtunion/wxapp/queryMediaOrderList?yodaReady=wx&csecappid=wxc32c3ddb81865d74&csecplatform=3&csecversionname=1.0.78&csecversion=1.3.0"

    payload = f'wm_actual_longitude=113.613319&wm_actual_latitude=34.748211&wm_longitude=113.613319&wm_latitude=34.748211&locatedCityId=73&cityId=73&orderViewId={orderViewId}&consumeCityName=%E5%85%A8%E9%83%A8&cityName=%E5%85%A8%E9%83%A8&pageSize=20&settleType=1&tenantType=10&pageNum=0&status=0&wm_logintoken={token}&userid={userid}&userId={userid}&user_id={userid}&lch=0&wm_uuid_source=server&wm_uuid=1191923790926557190&uuid=1191923790926557190&unionid=oNQu9t8NB_8JXj78m2GynFJJsRTo&open_id=orY-a7aLYtlG5amc3ZQFjafo56gw&openid=orY-a7aLYtlG5amc3ZQFjafo56gw&wm_appversion=1.0.78&wm_visitid=d8854a1e-dcfd-4bd3-8ba7-0a472b71eddc&wm_dplatform=windows&wm_dversion=4.1.2.17&wm_dtype=microsoft&wm_ctype=mtunion_wxapp&req_time=1761982422042&waimai_sign=%2F'
    headers = {
        'Accept': '*/*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Connection': 'keep-alive',
        'Referer': 'https://servicewechat.com/wxc32c3ddb81865d74/56/page-frame.html',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'cross-site',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36 MicroMessenger/7.0.20.1781(0x6700143B) NetType/WIFI MiniProgramEnv/Windows WindowsWechat/WMPF WindowsWechat(0x63090a13) UnifiedPCWindowsWechat(0xf2541211) XWEB/16815',
        'X-Requested-With': 'XMLHttpRequest',
        'csecuserid': f'{userid}',
        'csecuuid': '1191923790926557190',
        'geographyInfo': '%7B%7D',
        'mtgsig': '',
        'openId': 'orY-a7aLYtlG5amc3ZQFjafo56gw',
        'openIdCipher': 'AwQAAABJAgAAAAEAAAAyAAAAPLgC95WH3MyqngAoyM/hf1hEoKrGdo0pJ5DI44e1wGF9AT3PH7Wes03actC2n/GVnwfURonD78PewMUppAAAADhweG0THr5LNdyBO+Gisc2mbCbJrsXzdmfwo+5zl4NgWOwGFy2MeqnvSnzMR5xq7J9wxQz+lXlamA==',
        # AwQAAABJAgAAAAEAAAAyAAAAPLgC95WH3MyqngAoyM/hf1hEoKrGdo0pJ5DI44e1wGF9AT3PH7Wes03actC2n/GVnwfURonD78PewMUppAAAADink7sX5BzKp2vnJJUNS9wkAQiI/7e8SHBXxiMq9P5NO4XrA1XJNHq1q0sFh6uRE/kMq4cV1nzK2Q==
        # AwQAAABJAgAAAAEAAAAyAAAAPLgC95WH3MyqngAoyM/hf1hEoKrGdo0pJ5DI44e1wGF9AT3PH7Wes03actC2n/GVnwfURonD78PewMUppAAAADiOACSurP/mR2alu23ueBqU/6AOGfAVJ+bmsOv083Kbwb0ry5TOssADLygihPMNcjTiObnO/oa9mg==
        'swimlane': '',
        'token': f'{token}',
        'x-env': 'online',
        'xweb_xhr': '1',
        'Content-Type': 'application/x-www-form-urlencoded',
        'Host': 'media.meituan.com'
    }
    
    # 创建session以保持连接
    session = requests.Session()
    session.headers.update(headers)
    
    max_retries = 3  # 最大重试次数
    retry_count = 0
    
    while retry_count < max_retries:
        try:
            # 获取签名
            mtgsig = get_mtgsig_w_api("POST", url, payload, headers)
            session.headers.update({'mtgsig': mtgsig})
            
            print(f"[get_mt_order_rebate_info] 正在查询订单返利，尝试 {retry_count + 1}/{max_retries}")
            
            # 发起请求
            response = session.post(
                url,
                data=payload,
                timeout=30,  # 增加超时时间
                verify=False,
                allow_redirects=True
            )
            
            # 检查状态码
            if response.status_code != 200:
                print(f"[get_mt_order_rebate_info] HTTP状态码错误: {response.status_code}")
                raise Exception(f"HTTP状态码错误: {response.status_code}")
            
            # 尝试解析JSON
            try:
                res = response.json()
                
                # 检查是否遇到风控 (yodaCode: 406)
                if isinstance(res, dict) and res.get('yodaCode') == 406:
                    retry_count += 1
                    if retry_count < max_retries:
                        wait_time = retry_count * 1  # 每次等待时间递增
                        print(f"[get_mt_order_rebate_info] 检测到风控 (yodaCode: 406)，{wait_time}秒后重试 ({retry_count}/{max_retries})")
                        print(f"[get_mt_order_rebate_info] 风控信息: {res.get('msg', '未知')}")
                        time.sleep(wait_time)
                        continue
                    else:
                        print(f"[get_mt_order_rebate_info] 风控重试次数已用尽，返回风控响应")
                        session.close()
                        return res
                
                # 正常返回
                print(f"[get_mt_order_rebate_info] 查询成功")
                session.close()
                return res
            except json.JSONDecodeError:
                # 可能是文本响应（如403）
                res = response.text
                print(f"[get_mt_order_rebate_info] 返回文本: {res[:100]}")
                session.close()
                return res
            
        except (requests.exceptions.ConnectionError,
                requests.exceptions.Timeout,
                requests.exceptions.ReadTimeout,
                requests.exceptions.ChunkedEncodingError,
                ConnectionResetError) as e:
            retry_count += 1
            if retry_count < max_retries:
                wait_time = retry_count * 2  # 指数退避
                print(f"[get_mt_order_rebate_info] 连接错误，{wait_time}秒后重试 ({retry_count}/{max_retries}): {e}")
                time.sleep(wait_time)
            else:
                print(f"[get_mt_order_rebate_info] 达到最大重试次数，查询失败: {e}")
                session.close()
                return {"error": f"网络连接失败，已重试{max_retries}次: {str(e)}"}
        
        except Exception as e:
            print(f"[get_mt_order_rebate_info] 查询异常: {e}")
            session.close()
            return {"error": str(e)}
    
    session.close()
    return {"error": "查询失败，已达到最大重试次数"}


# ==================== 美团联盟接口函数 ====================
def generate_signature(
        http_method: str,
        path: str,
        query_params: dict,
        headers: dict,
        body: str,
        secret: str,
        app_key: str
) -> dict:
    """生成美团联盟API签名"""
    # 1. 生成时间戳
    timestamp = str(int(datetime.now().timestamp() * 1000))

    # 2. 计算Content-MD5
    content_md5 = ""
    if body:
        md5_hash = hashlib.md5(body.encode('utf-8')).digest()
        content_md5 = base64.b64encode(md5_hash).decode('utf-8')

    # 3. 组织Headers部分
    exclude_headers = ['S-Ca-Signature', 'S-Ca-Signature-Headers', 'Content-MD5']
    sign_headers = {k: v for k, v in headers.items() if k not in exclude_headers}

    # 确保必需的header
    sign_headers['S-Ca-Timestamp'] = timestamp
    sign_headers['S-Ca-App'] = app_key

    # 排序headers
    sorted_header_keys = sorted(sign_headers.keys())
    header_str = "".join(f"{k}:{sign_headers.get(k, '')}\n" for k in sorted_header_keys)

    # 4. 组织URL部分
    url_str = path
    if query_params:
        sorted_params = sorted(query_params.items(), key=lambda x: x[0])
        query_str = urlencode(sorted_params)
        url_str += f"?{query_str}"

    # 5. 构建待签名字符串
    string_to_sign = (
        f"{http_method.upper()}\n"
        f"{content_md5}\n"
        f"{header_str}"
        f"{url_str}"
    )

    # 6. 计算签名
    key_bytes = secret.encode('utf-8')
    message = string_to_sign.encode('utf-8')
    hmac_hash = hmac.new(key_bytes, message, hashlib.sha256).digest()
    signature = base64.b64encode(hmac_hash).decode('utf-8')
    print(f"Signature: {signature}")

    # 7. 准备headers
    signature_headers_list = ['S-Ca-Timestamp', 'S-Ca-App'] + sorted_header_keys
    signature_headers = {
        'S-Ca-App': app_key,
        'S-Ca-Signature': signature,
        'S-Ca-Timestamp': timestamp,
        'S-Ca-Signature-Headers': ",".join(signature_headers_list)
    }
    if http_method.upper() in ['POST', 'PUT'] and body:
        signature_headers['Content-MD5'] = content_md5

    return signature_headers

def query_orders(
        app_key: str,
        app_secret: str,
        query_params: Dict[str, Union[int, str, List[int], List[str]]],
        headers: Optional[Dict[str, str]] = None,
        base_url: str = "https://media.meituan.com/cps_open/common/api/v1/query_order"
) -> Dict:
    """查询订单函数"""
    # 准备请求数据
    http_method = "POST"
    path = "/cps_open/common/api/v1/query_order"
    body = json.dumps(query_params, ensure_ascii=False)

    # 合并headers
    request_headers = headers.copy() if headers else {}
    request_headers.update({
        "Content-Type": "application/json",
        "Accept": "application/json"
    })

    signature_headers = generate_signature(
        http_method=http_method,
        path=path,
        query_params={},  # 此API使用body参数而不是query参数
        headers=request_headers,
        body=body,
        secret=app_secret,
        app_key=app_key
    )

    # 添加签名头到请求头
    request_headers.update({
        k: v for k, v in signature_headers.items()
        if v  # 只添加有值的header
    })

    # 直接请求目标URL
    try:
        response = requests.post(
            base_url,
            headers=request_headers,
            data=body,
            timeout=30
        )

        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"\nRequest Failed: {str(e)}")
        if hasattr(e, 'response') and e.response:
            print(f"Error Response: {e.response.text}")
        raise Exception(f"Request failed: {str(e)}") from e

"""
# 查询参数
    params = {
        "limit": 20,
        "queryTimeType": 1,
        "page": 1,
        "platform": 2,
        "businessLine": [1, 2],
        "orderId": "5008031697162667606"
        # "startTime": int(datetime.now().timestamp()) - 86400,  # 24小时前
        # "endTime": int(datetime.now().timestamp())
    }

    # 调用接口
    try:
        result = query_orders(
            app_key=APP_KEY,
            app_secret=APP_SECRET,
            query_params=params
        )
        print("\n查询结果:")
        print(json.dumps(result, indent=2, ensure_ascii=False))
    except Exception as e:
        print(f"\n查询失败: {str(e)}")
"""


def query_products(
        app_key: str,
        app_secret: str,
        query_params: Dict[str, Union[int, str, List[int], List[str]]],
        headers: Optional[Dict[str, str]] = None,
        base_url: str = "https://media.meituan.com/cps_open/common/api/v1/query_coupon",
        max_retries: int = 10
) -> Dict:
    """
    查询美团售卖商品接口（带重试机制）

    :param app_key: 应用的key
    :param app_secret: 应用的secret
    :param query_params: 查询参数字典
    :param headers: 额外的请求头
    :param base_url: API基础地址
    :param max_retries: 最大重试次数
    :return: API响应数据
    """
    http_method = "POST"
    path = "/cps_open/common/api/v1/query_coupon"
    body = json.dumps(query_params, ensure_ascii=False)

    request_headers = headers.copy() if headers else {}
    request_headers.update({
        "Content-Type": "application/json",
        "Accept": "application/json"
    })

    # 生成签名
    signature_headers = generate_signature(
        http_method=http_method,
        path=path,
        query_params={},
        headers=request_headers,
        body=body,
        secret=app_secret,
        app_key=app_key
    )

    request_headers.update({
        k: v for k, v in signature_headers.items()
        if v
    })

    # 直接请求目标URL（带重试）
    last_error = None
    for retry in range(max_retries):
        try:
            response = requests.post(
                base_url,
                headers=request_headers,
                data=body,
                timeout=15
            )
            response.raise_for_status()
            result = response.json()
            
            if result and isinstance(result, dict):
                return result
            else:
                last_error = Exception(f"Invalid response data on attempt {retry + 1}")
                continue
                
        except requests.exceptions.RequestException as e:
            last_error = e
            if retry < max_retries - 1:
                continue
    
    raise Exception(f"Request failed after {max_retries} retries: {str(last_error)}") from last_error


def get_referral_link(
        app_key: str,
        app_secret: str,
        request_data: Dict[str, Union[int, str, List[int]]],
        headers: Optional[Dict[str, str]] = None,
        base_url: str = "https://media.meituan.com/cps_open/common/api/v1/get_referral_link",
        max_retries: int = 10
) -> Dict:
    """
    获取推广链接接口（带重试机制）

    :param app_key: 应用的key
    :param app_secret: 应用的secret
    :param request_data: 请求参数
    :param headers: 额外的请求头
    :param base_url: API基础地址
    :param max_retries: 最大重试次数
    :return: API响应数据
    """
    http_method = "POST"
    path = "/cps_open/common/api/v1/get_referral_link"
    body = json.dumps(request_data, ensure_ascii=False)

    request_headers = headers.copy() if headers else {}
    request_headers.update({
        "Content-Type": "application/json",
        "Accept": "application/json"
    })

    signature_headers = generate_signature(
        http_method=http_method,
        path=path,
        query_params={},
        headers=request_headers,
        body=body,
        secret=app_secret,
        app_key=app_key
    )

    request_headers.update({
        k: v for k, v in signature_headers.items()
        if v
    })

    # 直接请求目标URL（带重试）
    last_error = None
    for retry in range(max_retries):
        try:
            response = requests.post(
                base_url,
                headers=request_headers,
                data=body,
                timeout=15
            )
            response.raise_for_status()
            result = response.json()
            
            if result and isinstance(result, dict):
                return result
            else:
                last_error = Exception(f"Invalid response data on attempt {retry + 1}")
                continue
                
        except requests.exceptions.RequestException as e:
            last_error = e
            if retry < max_retries - 1:
                continue
    
    raise Exception(f"Request failed after {max_retries} retries: {str(last_error)}") from last_error


def get_sku_shop(limit, token, sku, offset, _safe_send_text, instance_wxid):
    """
    获取美团门店列表
    
    Args:
        limit: 每页数量
        token: 美团token
        sku: 商品ID（美团ID）
        offset: 偏移量
        
    Returns:
        list: 门店列表 [{"name": "门店名称", "cityName": "城市"}]
    """
    pois = []
    payload = {}
    headers = {
        'Host': 'apimobile.meituan.com',
        'Connection': 'keep-alive',
        'Accept': 'application/json, text/plain, */*',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 MicroMessenger/7.0.20.1781(0x6700143B) NetType/WIFI MiniProgramEnv/Windows WindowsWechat/WMPF WindowsWechat(0x63090a13)XWEB/14315',
        'Origin': 'https://awp.meituan.com',
        'Sec-Fetch-Site': 'same-site',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Dest': 'empty',
        'Referer': 'https://awp.meituan.com/',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
    }
    
    # 创建session以保持连接
    session = requests.Session()
    session.headers.update(headers)
    
    page = 1
    max_pages = 100  # 最大页数限制，防止无限循环
    max_retries = 3  # 每个请求最大重试次数
    
    while page <= max_pages:
        retry_count = 0
        success = False
        server_index = 0  # 服务器索引：0=首次, 1=第一次重试, 2=第二次重试
        
        while retry_count < max_retries and not success:
            try:
                url = f"https://apimobile.meituan.com/group/v2/deal/{sku}/branches?token={token}&preCityId=1&offset={offset}&limit={limit}&platform=mtapp&os=android&dpId=&chooseCity=0&chooseAllCity=0&bundle_version=1.23.0&source=order&yodaReady=h5&csecplatform=4&csecversion=4.0.3"
                params = extract_url_params(url)
                
                # 根据server_index获取签名URL（每次重试使用不同的服务器）
                url2 = get_mtgsig_url_get(url, params, server_index)
                
                # 直接请求美团接口
                response = session.get(
                    url2, 
                    data=payload, 
                    timeout=30,
                    verify=False,  # 跳过SSL证书验证（解决打包后证书问题）
                    allow_redirects=True
                )
                
                # 检测403风控
                if response.status_code == 403 or '403 Forbidden' in response.text:
                    # 如果遇到403，最多重试2次，每次使用不同的服务器
                    if server_index < 2:
                        server_index += 1
                        print(f"[get_sku_shop] 遇到403风控，尝试使用服务器[{server_index}]重新获取签名...")
                        time.sleep(1)
                        continue  # 继续循环，使用新的server_index
                    else:
                        # 已经重试2次仍然403，放弃
                        raise Exception(f"抓取失败，请稍后再试")
                
                # 检查状态码
                if response.status_code != 200:
                    raise Exception(f"HTTP状态码错误: {response.status_code}")
                
                # 尝试解析JSON
                try:
                    res = response.json()
                except json.JSONDecodeError as e:
                    raise Exception(f"解析响应失败，已抓取{len(pois)}家门店，响应内容: {response.text[:200]}")
                
                # 检测风控
                if res.get('yodaCode') == 406 or '您的网络好像不太给力' in res.get('msg', ''):
                    general_page_url = res.get('customData', {}).get('generalPageUrl', '')
                    print(f"[get_sku_shop] 检测到风控，yodaCode={res.get('yodaCode')}, msg={res}")
                    notification_msg = f"🚨 风控通知\n风控验证URL:\n{res}"
                    admin_wxid = "wxid_3intiqznkov222"
                    _safe_send_text(instance_wxid, admin_wxid, notification_msg)
                    raise MeituanRiskControlException(
                        f"遇到美团风控: {res.get('msg', '未知错误')}",
                        general_page_url=general_page_url
                    )
                
                try:
                    data = res.get('data', [])
                except Exception as e:
                    data = []
                
                if not data:
                    # 没有更多数据，关闭session并返回结果
                    print(f"[get_sku_shop] 没有更多数据，抓取完成，共 {len(pois)} 家门店")
                    session.close()
                    return pois
                else:
                    for store in data:
                        name = store.get('name', '')
                        cityName = store.get('cityName', '')
                        store_dict = {
                            'name': name,
                            'cityName': cityName
                        }
                        pois.append(store_dict)
                    success = True
                    
                    # 增加请求间隔，避免请求过快
                    time.sleep(1)
                    offset += limit
                    page += 1
                
            except MeituanRiskControlException:
                # 风控异常必须向上抛出，不返回部分数据
                session.close()
                raise
            
            except (requests.exceptions.ConnectionError, 
                    requests.exceptions.ChunkedEncodingError,
                    ConnectionResetError) as e:
                retry_count += 1
                if retry_count < max_retries:
                    wait_time = retry_count * 2  # 指数退避
                    print(f"连接错误，{wait_time}秒后重试 ({retry_count}/{max_retries}): {e}")
                    time.sleep(wait_time)
                else:
                    # 网络错误重试失败，必须抛出异常
                    session.close()
                    raise Exception(f"网络连接失败，已重试{max_retries}次: {str(e)}")
                        
            except Exception as e:
                # 其他异常（如403错误等）也必须抛出，不返回部分数据
                session.close()
                raise e
        
        # 如果重试后仍未成功，退出循环
        if not success:
            break
    
    session.close()
    return pois


# ==================== botTools集成功能 ====================

def process_authorization_db(bot_wxid, message, final_from_wxid, send_callback):
    """处理授权命令"""
    # 解析消息类型和授权码
    parts = message.split('@')
    auth_type = parts[0]
    old_setid = parts[1] if len(parts) > 1 else None
    new_setid = parts[2] if len(parts) > 2 else None

    # 连接数据库
    if not PYMYSQL_AVAILABLE:
        send_callback(bot_wxid, final_from_wxid, "系统错误：缺少pymysql模块")
        return
        
    try:
        conn = pymysql.connect(
            host='115.190.182.82',
            user='root',
            password='ltx123589.',
            database='userSetId',
            charset='utf8mb4'
        )
    except Exception as e:
        print(f"数据库连接失败: {e}")
        return

    try:
        with conn.cursor() as cursor:
            # 正式授权/试用授权/天卡授权
            if auth_type in ["正式授权", "试用授权", "天卡授权"]:
                setid = old_setid

                # 检查setId是否存在
                cursor.execute("SELECT COUNT(*) FROM setid_new WHERE setId = %s", (setid,))
                exists = cursor.fetchone()[0] > 0

                # 计算新时间
                now = datetime.now()
                if auth_type == "正式授权":
                    # 下个月的今天（处理跨年和月份天数问题）
                    next_month = now.month % 12 + 1
                    year = now.year + (1 if next_month == 1 else 0)

                    # 处理月份天数不足的情况（如1月31日->2月28/29日）
                    try:
                        new_time = now.replace(year=year, month=next_month)
                    except ValueError:
                        # 如果目标月份没有当前日期的天数，使用目标月份的最后一天
                        # 例如：1月31日 -> 2月28日（或29日）
                        if next_month == 12:
                            # 下个月是12月，下下个月是次年1月
                            next_next_month = 1
                            next_year = year + 1
                        else:
                            next_next_month = next_month + 1
                            next_year = year

                        # 获取目标月份的最后一天
                        last_day_of_month = (datetime(next_year, next_next_month, 1) - timedelta(days=1)).day
                        new_time = now.replace(year=year, month=next_month, day=last_day_of_month)
                elif auth_type == "试用授权":
                    new_time = now + timedelta(days=3)
                else:  # 天卡授权
                    new_time = now + timedelta(days=1)

                # 更新或插入记录
                if exists:
                    cursor.execute(
                        "UPDATE setid_new SET userendtime = %s WHERE setId = %s",
                        (new_time, setid)
                    )
                else:
                    cursor.execute(
                        "INSERT INTO setid_new (setId, userendtime, user) VALUES (%s, %s, %s)",
                        (setid, new_time, setid)
                    )
                conn.commit()
                send_callback(bot_wxid, final_from_wxid, f"{setid} 授权更新成功")

            # 更换授权
            elif auth_type == "更换授权" and new_setid:
                # 检查原setId是否存在
                cursor.execute("SELECT COUNT(*) FROM setid_new WHERE setId = %s", (old_setid,))
                if cursor.fetchone()[0] == 0:
                    send_callback(bot_wxid, final_from_wxid, f"{old_setid} 不存在")
                    return

                # 更新setId和user字段
                cursor.execute(
                    "UPDATE setid_new SET setId = %s, user = %s WHERE setId = %s",
                    (new_setid, new_setid, old_setid)
                )
                conn.commit()
                send_callback(bot_wxid, final_from_wxid, f"授权已从 {old_setid} 更换为 {new_setid}")

            # 删除授权
            elif auth_type == "删除授权":
                if len(parts) < 2:
                    return
                setid = parts[1]

                # 检查setId是否存在
                cursor.execute("SELECT COUNT(*) FROM setid_new WHERE setId = %s", (setid,))
                if cursor.fetchone()[0] == 0:
                    send_callback(bot_wxid, final_from_wxid, f"{setid} 不存在")
                    return

                # 删除记录
                cursor.execute("DELETE FROM setid_new WHERE setId = %s", (setid,))
                conn.commit()
                send_callback(bot_wxid, final_from_wxid, f"{setid} 已成功删除")

    except Exception as e:
        conn.rollback()
        print(f"处理授权出错: {str(e)}")
    finally:
        conn.close()


def process_peer_authorization_db(bot_wxid, message, final_from_wxid, send_callback):
    """处理同行授权命令"""
    # 解析消息类型和授权码
    parts = message.split('@')
    auth_type = parts[0]
    setid = parts[1] if len(parts) > 1 else None
    username = parts[2] if len(parts) > 2 else None

    # 连接数据库
    if not PYMYSQL_AVAILABLE:
        send_callback(bot_wxid, final_from_wxid, "系统错误：缺少pymysql模块")
        return

    try:
        conn = pymysql.connect(
            host='115.190.182.82',
            user='root',
            password='ltx123589.',
            database='userSetId',
            charset='utf8mb4'
        )
    except Exception as e:
        print(f"同行授权数据库连接失败: {e}")
        return

    try:
        with conn.cursor() as cursor:
            # 同行授权
            if auth_type == "同行授权" and setid and username:
                # 检查setId是否存在
                cursor.execute("SELECT COUNT(*) FROM setid_out WHERE setId = %s", (setid,))
                exists = cursor.fetchone()[0] > 0

                # 计算新时间：今天+30天
                now = datetime.now()
                new_time = now + timedelta(days=30)

                # 更新或插入记录
                if exists:
                    cursor.execute(
                        "UPDATE setid_out SET userendtime = %s, user = %s WHERE setId = %s",
                        (new_time, username, setid)
                    )
                    send_callback(bot_wxid, final_from_wxid, f"同行授权更新成功\nsetId: {setid}\n用户: {username}\n到期时间: {new_time.strftime('%Y-%m-%d %H:%M:%S')}")
                else:
                    cursor.execute(
                        "INSERT INTO setid_out (setId, userendtime, user) VALUES (%s, %s, %s)",
                        (setid, new_time, username)
                    )
                    send_callback(bot_wxid, final_from_wxid, f"同行授权添加成功\nsetId: {setid}\n用户: {username}\n到期时间: {new_time.strftime('%Y-%m-%d %H:%M:%S')}")

                conn.commit()

            # 删除同行授权
            elif auth_type == "删除同行授权" and setid:
                # 检查setId是否存在
                cursor.execute("SELECT COUNT(*) FROM setid_out WHERE setId = %s", (setid,))
                if cursor.fetchone()[0] == 0:
                    send_callback(bot_wxid, final_from_wxid, f"同行授权 {setid} 不存在")
                    return

                # 删除记录
                cursor.execute("DELETE FROM setid_out WHERE setId = %s", (setid,))
                conn.commit()
                send_callback(bot_wxid, final_from_wxid, f"同行授权 {setid} 已成功删除")

    except Exception as e:
        conn.rollback()
        print(f"处理同行授权出错: {str(e)}")
        send_callback(bot_wxid, final_from_wxid, f"同行授权处理失败: {str(e)}")
    finally:
        conn.close()


class MultiFunctionPlugin(PluginBase):
    """聚合bot插件 - 关键词回复、用户管理、日志、授权管理等"""

    def __init__(self):
        super().__init__()
        # 关键词回复数据
        self.keywords = []  # [{"keyword": "关键词", "reply": "回复话术"}]

        # 用户管理数据
        self.users = []  # [{"nickname": "昵称", "wxid": "wxid", "open_time": "开户时间", "balance": 余额, "status": "状态"}]

        # 日志数据
        self.logs = []  # [{"time": "时间", "level": "级别", "message": "消息"}]

        # 美团抓取配置
        self.mt_token = self.get_config("mt_token", "AgGSJNiBiU5m6N1FLRE36DW_mK7DpQINgEb4H87pnv2o9i21TExm9qoPto8sx7rMrkJO3")
        self.mt_limit = self.get_config("mt_limit", 50)  # 每页数量
        self.delete_after_send = self.get_config("delete_after_send", False)  # 发送后是否删除文件

        # 美团联盟配置
        self.mt_app_key = self.get_config("mt_app_key", "28acf264f2354039a4c113eb8f95b69d")
        self.mt_app_secret = self.get_config("mt_app_secret", "b027653fb0e1432cb219faa93cdafeab")

        # 用户会话管理（记录用户当前的操作状态）
        self.user_sessions = {}  # {wxid: {"type": "mt_menu", "mt_id": "xxx", "content": "xxx", "instance_wxid": "xxx"}}

        # 日志保存计数器（每添加50条日志保存一次）
        self.log_save_counter = 0

        # 城市字典（先初始化为空，后面加载）
        self.city_dict = {}

        # 创建Excel文件保存目录
        self.excel_output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "excel_output")
        if not os.path.exists(self.excel_output_dir):
            try:
                os.makedirs(self.excel_output_dir)
            except:
                # 如果创建失败，使用当前目录
                self.excel_output_dir = os.getcwd()

        # 创建日志文件保存目录
        self.log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
        if not os.path.exists(self.log_dir):
            try:
                os.makedirs(self.log_dir)
            except:
                self.log_dir = os.path.dirname(os.path.abspath(__file__))

        # 日志文件路径
        self.log_file = os.path.join(self.log_dir, "聚合bot_logs.json")

        # 会员数据文件路径
        self.vip_file = os.path.join(self.log_dir, "vip_members.json")

        # 会员数据
        self.vip_members = {}  # {wxid: {"type": "月卡/季卡/年卡", "expire_time": "2025-12-31 23:59:59"}}

        # 管理员数据文件路径
        self.admin_file = os.path.join(self.log_dir, "admins.json")

        # 超级管理员（固定）
        self.super_admin = "wxid_3intiqznkov222"

        # 管理员列表
        self.admins = []  # [wxid1, wxid2, ...]

        # GM功能允许的群列表
        self.allowed_gm_groups = [
            "50037556302@chatroom", "44370239233@chatroom", "48971005641@chatroom",
            "48253802651@chatroom", "49769395441@chatroom", "43574753937@chatroom",
            "47282914754@chatroom", "46123714680@chatroom"
        ]

        # 卡密管理数据文件路径
        self.card_key_file = os.path.join(self.log_dir, "card_keys.json")

        # 卡密数据
        self.card_keys = []  # [{"card_key": "JHKM+16位MD5", "amount": 金额, "create_time": "创建时间", "used": False, "used_time": None, "used_by": None}]

        # 企业配置文件路径
        self.enterprise_config_file = os.path.join(self.log_dir, "enterprise_config.json")

        # 企业配置数据
        self.enterprise_configs = {}  # {sk: {"app_key": "xxx", "app_secret": "xxx", "name": "企业名称"}}

    def get_plugin_info(self):
        """获取插件信息"""
        return {
            "name": "聚合bot",
            "version": "1.0.0",
            "author": "问世科技",
            "description": "提供关键词回复、用户管理、日志记录、授权管理、券码生成、闲管家卡密、作图等功能"
        }

    def _load_logs_from_file(self):
        """从文件加载日志"""
        try:
            if os.path.exists(self.log_file):
                with open(self.log_file, 'r', encoding='utf-8') as f:
                    logs = json.load(f)
                    # 限制日志数量，只保留最近的1000条
                    if len(logs) > 1000:
                        logs = logs[-1000:]
                    return logs
        except Exception as e:
            self.log(f"从文件加载日志失败: {e}")
        return []

    def _load_vip_members(self):
        """从文件加载会员数据"""
        try:
            if os.path.exists(self.vip_file):
                with open(self.vip_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            self.log(f"从文件加载会员数据失败: {e}")
        return {}

    def _save_vip_members(self):
        """保存会员数据到文件"""
        try:
            with open(self.vip_file, 'w', encoding='utf-8') as f:
                json.dump(self.vip_members, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            self.log(f"保存会员数据到文件失败: {e}")
            return False

    def _load_admins(self):
        """从文件加载管理员列表"""
        try:
            if os.path.exists(self.admin_file):
                with open(self.admin_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            self.log(f"从文件加载管理员列表失败: {e}")
        return []

    def _save_admins(self):
        """保存管理员列表到文件"""
        try:
            with open(self.admin_file, 'w', encoding='utf-8') as f:
                json.dump(self.admins, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            self.log(f"保存管理员列表到文件失败: {e}")
            return False

    def _load_card_keys(self):
        """从文件加载卡密数据"""
        try:
            if os.path.exists(self.card_key_file):
                with open(self.card_key_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            self.log(f"从文件加载卡密数据失败: {e}")
        return []

    def _save_card_keys(self):
        """保存卡密数据到文件"""
        try:
            with open(self.card_key_file, 'w', encoding='utf-8') as f:
                json.dump(self.card_keys, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            self.log(f"保存卡密数据到文件失败: {e}")
            return False

    def _load_enterprise_configs(self):
        """从文件加载企业配置数据"""
        try:
            if os.path.exists(self.enterprise_config_file):
                with open(self.enterprise_config_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                # 创建默认配置（包含问世科技）
                default_config = {
                    "f08qeT7qae13We76XyaorQQ9TalPx4Oi1almqe4rXFbNey53": {
                        "app_key": "28acf264f2354039a4c113eb8f95b69d",
                        "app_secret": "b027653fb0e1432cb219faa93cdafeab",
                        "name": "问世科技"
                    },
                    "g39sfV9scg35Yf98ZacqtsSS1VcnRz6Qk3cmosg6tZHdPg07": {
                        "app_key": "b24e381b3ded4e288be48ad6bcd9a5dc",
                        "app_secret": "580983ee83a440fcafb3358a956fc4cd",
                        "name": "刺猬寄售"
                  }
                }
                # 保存默认配置
                with open(self.enterprise_config_file, 'w', encoding='utf-8') as f:
                    json.dump(default_config, f, ensure_ascii=False, indent=2)
                return default_config
        except Exception as e:
            self.log(f"从文件加载企业配置失败: {e}")
            # 返回默认配置
            return {
                "f08qeT7qae13We76XyaorQQ9TalPx4Oi1almqe4rXFbNey53": {
                    "app_key": "28acf264f2354039a4c113eb8f95b69d",
                    "app_secret": "b027653fb0e1432cb219faa93cdafeab",
                    "name": "问世科技"
                }
            }
    
    def _save_enterprise_configs(self):
        """保存企业配置数据到文件"""
        try:
            with open(self.enterprise_config_file, 'w', encoding='utf-8') as f:
                json.dump(self.enterprise_configs, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            self.log(f"保存企业配置到文件失败: {e}")
            return False
    
    def _generate_card_key(self, amount, count=1):
        """
        生成指定面额的卡密
        
        Args:
            amount: 面额（￥）
            count: 生成数量，默认1
            
        Returns:
            list: 生成的卡密列表 ["JHKM+16位MD5", ...]
        """
        generated_keys = []
        
        for _ in range(count):
            # 生成唯一的随机字符串（时间戳+随机数）
            unique_str = f"{time.time()}{os.urandom(16).hex()}"
            
            # 计算MD5并取前16位
            md5_hash = hashlib.md5(unique_str.encode()).hexdigest()[:16]
            
            # 组合卡密：JHKM + 16位MD5
            card_key = f"JHKM{md5_hash}"
            
            # 保存卡密信息
            card_info = {
                "card_key": card_key,
                "amount": amount,
                "create_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "used": False,
                "used_time": None,
                "used_by": None
            }
            
            self.card_keys.append(card_info)
            generated_keys.append(card_key)
            
            # 短暂延迟，确保时间戳不同
            if count > 1:
                time.sleep(0.001)
        
        # 保存到文件
        self._save_card_keys()
        
        return generated_keys
    
    def _use_card_key(self, card_key, wxid):
        """
        使用卡密充值
        
        Args:
            card_key: 卡密
            wxid: 使用者WXID
            
        Returns:
            tuple: (成功标志, 消息文本, 充值金额)
        """
        # 查找卡密
        for card_info in self.card_keys:
            if card_info["card_key"] == card_key:
                # 检查是否已使用
                if card_info["used"]:
                    return False, f"❌ 卡密已被使用\n使用时间：{card_info['used_time']}", 0
                
                # 标记为已使用
                card_info["used"] = True
                card_info["used_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                card_info["used_by"] = wxid
                
                amount = card_info["amount"]
                
                # 查找用户
                user = None
                for u in self.users:
                    if u["wxid"] == wxid:
                        user = u
                        break
                
                # 如果用户不存在，提示先开户
                if not user:
                    return False, "❌ 您还未开户，请先发送\"聚合开户\"进行注册\n\n发送\"教程\"查看使用教程", 0
                
                # 增加用户余额
                user["balance"] += amount
                
                # 保存数据
                self._save_card_keys()
                self.set_config("users", self.users)
                
                # 记录日志
                self.add_log("INFO", f"用户 {wxid} 使用卡密 {card_key} 充值 {amount} ￥")
                
                return True, f"✅ 充值成功！\n充值金额：{amount}￥\n当前余额：{user['balance']}￥", amount
        
        # 卡密不存在
        return False, "❌ 卡密不存在或已失效", 0
    
    def _save_logs_to_file(self):
        """保存日志到文件"""
        try:
            # 限制日志数量，只保留最近的1000条
            logs_to_save = self.logs[-1000:] if len(self.logs) > 1000 else self.logs
            
            with open(self.log_file, 'w', encoding='utf-8') as f:
                json.dump(logs_to_save, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            self.log(f"保存日志到文件失败: {e}")
            return False
    
    def on_load(self):
        """插件加载时调用"""
        self.log("聚合bot插件正在加载...")
        
        # 加载配置
        try:
            self.keywords = self.get_config("keywords", [])
            self.users = self.get_config("users", [])
            
            # 从文件加载日志
            self.logs = self._load_logs_from_file()
            
            # 从文件加载会员数据
            self.vip_members = self._load_vip_members()
            
            # 从文件加载管理员列表
            self.admins = self._load_admins()
            
            # 从文件加载卡密数据
            self.card_keys = self._load_card_keys()
            
            # 从文件加载企业配置
            self.enterprise_configs = self._load_enterprise_configs()
            
            # 加载城市字典
            self._load_city_dict()
            
            # 如果文件中没有日志，尝试从配置文件迁移
            if not self.logs:
                old_logs = self.get_config("logs", [])
                if old_logs:
                    self.log(f"检测到配置文件中的日志数据，正在迁移到文件...")
                    self.logs = old_logs
                    self._save_logs_to_file()
                    # 清空配置文件中的日志，避免重复
                    self.set_config("logs", [])
                    self.log("日志数据已迁移到文件")
            
            # 如果当前配置为空，尝试从旧配置迁移数据（兼容性处理）
            if not self.users:
                # 尝试从旧的"多功能插件"配置加载
                old_users = self._bridge.main_window.config.get("plugin_多功能插件_users", [])
                if old_users:
                    self.log(f"检测到旧配置数据，正在迁移 {len(old_users)} 个用户...")
                    self.users = old_users
                    self.set_config("users", self.users)  # 保存到新配置键
                    self.log("用户数据迁移完成")
            
            if not self.keywords:
                old_keywords = self._bridge.main_window.config.get("plugin_多功能插件_keywords", [])
                if old_keywords:
                    self.log(f"检测到旧配置数据，正在迁移 {len(old_keywords)} 个关键词...")
                    self.keywords = old_keywords
                    self.set_config("keywords", self.keywords)
                    self.log("关键词数据迁移完成")
            
            # 记录加载的数据量
            self.log(f"已加载 {len(self.keywords)} 个关键词")
            self.log(f"已加载 {len(self.users)} 个用户")
            self.log(f"已从文件加载 {len(self.logs)} 条日志")
            
            # 显示用户详细信息（用于调试）
            if self.users:
                self.log(f"用户列表: {[u['wxid'] for u in self.users]}")
            
            self.log("聚合bot插件加载成功")
            return True
            
        except Exception as e:
            import traceback
            self.log(f"加载配置时出错: {e}\n{traceback.format_exc()}")
            # 即使出错也返回True，使用默认空数据
            self.keywords = []
            self.users = []
            self.logs = []
            return True
    
    def on_unload(self):
        """插件卸载时调用"""
        self.log("聚合bot插件正在卸载...")
        
        # 保存配置（不包括日志）
        self.set_config("keywords", self.keywords)
        self.set_config("users", self.users)
        
        # 保存日志到文件
        self._save_logs_to_file()
        
        self.log("聚合bot插件已卸载")
    
    def on_enable(self):
        """插件启用时调用"""
        super().on_enable()
        self.add_log("INFO", "聚合bot插件已启用")
    
    def on_disable(self):
        """插件禁用时调用"""
        super().on_disable()
        
        # 禁用时也保存配置，防止数据丢失（不包括日志）
        self.set_config("keywords", self.keywords)
        self.set_config("users", self.users)
        
        # 保存日志到文件
        self._save_logs_to_file()
        
        self.add_log("INFO", "聚合bot插件已禁用")
    
    def on_message(self, instance_wxid, msg_data):
        """收到消息时调用"""
        if not self.is_enabled():
            return
        
        try:
            msg_type = msg_data.get("msg_type", 0)
            content = msg_data.get("content", "")
            from_wxid = msg_data.get("wxid1", "")
            wxid2 = msg_data.get("wxid2", "")
            is_self = msg_data.get("is_self", 0)
            
            # 忽略自己发送的消息（msg_type=51的特殊处理）
            if is_self == 1 and msg_type == 51:
                return
            
            # 判断消息来源：群聊或私聊
            if from_wxid and from_wxid.endswith('@chatroom'):
                final_from_wxid = from_wxid  # 群聊ID
                is_group = True
            else:
                # 私聊消息：如果是自己发的(is_self=1)，来源是wxid2；否则是from_wxid
                final_from_wxid = wxid2 if is_self == 1 else from_wxid
                is_group = False
            
            # 忽略自己发送的其他消息
            if is_self == 1:
                return
            
            # ==================== 处理各种功能 ====================
            
            # 1. 处理文本消息（仅私聊）
            if msg_type == 1 and not is_group:
                # 1.1 处理聚合开户命令
                if content.strip() == "聚合开户":
                    self._handle_register_user(instance_wxid, from_wxid)
                    return
                
                # 1.2 检查是否有待处理的会话（用户选择）
                if from_wxid in self.user_sessions:
                    # 检查用户是否已开户
                    has_permission, error_msg, user = self._check_user_permission(from_wxid)
                    if not has_permission:
                        self._safe_send_text(instance_wxid, from_wxid, error_msg)
                        # 清除会话
                        if from_wxid in self.user_sessions:
                            del self.user_sessions[from_wxid]
                        return
                    self._handle_user_choice(instance_wxid, from_wxid, content.strip())
                    return
                
                # 1.3 处理"我的wxid"命令
                if "我的wxid" in content:
                    self._safe_send_text(instance_wxid, final_from_wxid, final_from_wxid)
                    return
                
                # 1.4 处理"教程"命令
                if content.strip() == "教程":
                    self._safe_send_text(instance_wxid, from_wxid, "【腾讯文档】聚合机器人使用教程\nhttps://docs.qq.com/doc/DSVdBWnVYb1pzU2Np")
                    return
                
                # 1.4.1 处理"余额"命令
                if content.strip() == "余额":
                    self._handle_check_balance(instance_wxid, from_wxid)
                    return
                
                # 1.5 处理"查返利"命令
                if content.strip() == "查返利":
                    # 检查用户是否已开户
                    has_permission, error_msg, user = self._check_user_permission(from_wxid)
                    if not has_permission:
                        self._safe_send_text(instance_wxid, from_wxid, error_msg)
                        return
                    self._handle_check_rebate(instance_wxid, from_wxid)
                    return
                
                # 1.5.0.1 处理"查企业返利"命令
                if content.strip() == "查企业返利":
                    # 检查用户是否已开户
                    has_permission, error_msg, user = self._check_user_permission(from_wxid)
                    if not has_permission:
                        self._safe_send_text(instance_wxid, from_wxid, error_msg)
                        return
                    self._handle_check_enterprise_rebate(instance_wxid, from_wxid)
                    return
                
                # 1.5.1 处理"联盟榜单"命令
                if content.strip() == "联盟榜单":
                    # 检查用户是否已开户
                    has_permission, error_msg, user = self._check_user_permission(from_wxid)
                    if not has_permission:
                        self._safe_send_text(instance_wxid, from_wxid, error_msg)
                        return
                    if hasattr(self, '_handle_alliance_ranking'):
                        self._handle_alliance_ranking(instance_wxid, from_wxid)
                    else:
                        self._safe_send_text(instance_wxid, from_wxid, "⚠️ 联盟榜单功能暂未开放")
                    return
                
                # 1.6 处理"作图"命令
                if content.startswith("作图"):
                    # 检查用户是否已开户
                    has_permission, error_msg, user = self._check_user_permission(from_wxid)
                    if not has_permission:
                        self._safe_send_text(instance_wxid, from_wxid, error_msg)
                        return
                    self._handle_manual_coupon(instance_wxid, from_wxid, content)
                    return
                
                # 1.7 处理闲管家卡密链接
                if "m.goofish.pro/kami/" in content or "api.goofish.pro" in content:
                    # 检查用户是否已开户
                    has_permission, error_msg, user = self._check_user_permission(from_wxid)
                    if not has_permission:
                        self._safe_send_text(instance_wxid, from_wxid, error_msg)
                        return
                    self._handle_goofish_kami(instance_wxid, from_wxid, content)
                    return
                
                # 1.8 处理91卡券链接
                if "mai.91kami.com/cpd/" in content or "mai.91kami.com" in content:
                    # 检查用户是否已开户
                    has_permission, error_msg, user = self._check_user_permission(from_wxid)
                    if not has_permission:
                        self._safe_send_text(instance_wxid, from_wxid, error_msg)
                        return
                    self._handle_91kami(instance_wxid, from_wxid, content)
                    return
                
                # 1.9 处理卡密充值（用户发送JHKM开头的卡密）
                if content.strip().startswith("JHKM") and len(content.strip()) == 20:  # JHKM + 16位MD5
                    self._handle_card_key_recharge(instance_wxid, from_wxid, content.strip())
                    return
            
            # 2. 处理管理员命令（仅私聊）
            if msg_type == 1 and not is_group:
                # 处理设置管理员命令（仅超级管理员）
                if content.startswith("设置管理员@"):
                    self._handle_set_admin(instance_wxid, from_wxid, content)
                    return
                
                # 处理查看管理员列表命令（管理员和超级管理员可用）
                if content.strip() == "管理员列表":
                    is_admin, admin_type = self._check_admin_permission(from_wxid)
                    if not is_admin:
                        self._safe_send_text(instance_wxid, from_wxid, "❌ 权限不足，只有管理员和超级管理员可以查看管理员列表")
                        return
                    
                    admin_list = "\n".join(self.admins) if self.admins else "暂无普通管理员"
                    msg = f"👥 管理员列表\n\n超级管理员:\n{self.super_admin}\n\n普通管理员（{len(self.admins)}人）:\n{admin_list}"
                    self._safe_send_text(instance_wxid, from_wxid, msg)
                    return
            
            # 3. 处理授权命令（仅私聊，需要管理员权限）
            if msg_type == 1 and not is_group:
                # 检查管理员权限
                is_admin, admin_type = self._check_admin_permission(from_wxid)
                
                if content.startswith("正式授权@") or content.startswith("试用授权@") or \
                   content.startswith("天卡授权@") or content.startswith("更换授权@") or \
                   content.startswith("删除授权@"):
                    if not is_admin:
                        self._safe_send_text(instance_wxid, from_wxid, "❌ 权限不足，只有管理员和超级管理员可以执行授权操作")
                        return
                    process_authorization_db(instance_wxid, content, final_from_wxid, self._safe_send_text)
                    return
                
                # 处理同行授权
                if content.startswith("同行授权@") or content.startswith("删除同行授权@"):
                    if not is_admin:
                        self._safe_send_text(instance_wxid, from_wxid, "❌ 权限不足，只有管理员和超级管理员可以执行授权操作")
                        return
                    process_peer_authorization_db(instance_wxid, content, final_from_wxid, self._safe_send_text)
                    return
                
                # 处理会员授权（月卡/季卡/年卡）
                if content.startswith("月卡授权@") or content.startswith("季卡授权@") or content.startswith("年卡授权@"):
                    if not is_admin:
                        self._safe_send_text(instance_wxid, from_wxid, "❌ 权限不足，只有管理员和超级管理员可以执行授权操作")
                        return
                    self._handle_vip_authorization(instance_wxid, from_wxid, content)
                    return
                
                # 处理管理员生成卡密命令
                if content.startswith("生成卡密@"):
                    if not is_admin:
                        self._safe_send_text(instance_wxid, from_wxid, "❌ 权限不足，只有管理员和超级管理员可以生成卡密")
                        return
                    self._handle_admin_generate_card_key(instance_wxid, from_wxid, content)
                    return
            
            # 4. 处理GM功能（支持私聊和特定群组）
            if msg_type == 1 and "商家撤销了您的券码" in content:
                # 私聊直接处理，群聊需要在允许列表中
                if not is_group or final_from_wxid in self.allowed_gm_groups:
                    # 私聊时检查用户是否已开户
                    if not is_group:
                        has_permission, error_msg, user = self._check_user_permission(from_wxid)
                        if not has_permission:
                            self._safe_send_text(instance_wxid, from_wxid, error_msg)
                            return
                    self._handle_gm_coupon(instance_wxid, final_from_wxid, content)
                    return
            
            # 5. 处理小程序链接消息（仅私聊）
            if msg_type == 49 and not is_group:
                # 检查用户是否已开户
                has_permission, error_msg, user = self._check_user_permission(from_wxid)
                if not has_permission:
                    self._safe_send_text(instance_wxid, from_wxid, error_msg)
                    return
                self._handle_miniprogram_message(instance_wxid, from_wxid, content)
            
            # 这里可以实现关键词匹配自动回复
            # TODO: 实现关键词自动回复功能
        
        except Exception as e:
            self.add_log("ERROR", f"处理消息时出错: {e}")
            import traceback
            self.add_log("ERROR", f"详细错误: {traceback.format_exc()}")
    
    def add_log(self, level, message):
        """添加日志"""
        log_entry = {
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "level": level,
            "message": message
        }
        self.logs.append(log_entry)
        
        # 限制日志数量
        if len(self.logs) > 1000:
            self.logs.pop(0)
        
        # 定期保存日志到文件（每50条保存一次）
        self.log_save_counter += 1
        if self.log_save_counter >= 50:
            self._save_logs_to_file()
            self.log_save_counter = 0
        
        # 使用QTimer在主线程中更新UI（避免跨线程UI操作崩溃）
        self._safe_update_log_table()
    
    def _safe_update_log_table(self):
        """安全更新日志表格（线程安全）"""
        try:
            if self._ui_widget and hasattr(self, 'log_table'):
                QTimer.singleShot(0, self._update_log_table)
            else:
                # UI还未创建，稍后会自动加载
                pass
        except Exception as e:
            # 避免在日志更新中记录日志导致循环
            import traceback
            print(f"调度日志UI更新失败: {e}\n{traceback.format_exc()}")
    
    def create_ui(self):
        """创建插件UI"""
        widget = QWidget()
        layout = QVBoxLayout()
        widget.setLayout(layout)
        
        # 标题
        title = QLabel("🎯 多功能插件控制面板")
        title.setFont(QFont("微软雅黑", 14, QFont.Bold))
        title.setStyleSheet("color: #1890ff; padding: 10px;")
        layout.addWidget(title)
        
        # 创建标签页
        self.tab_widget = QTabWidget()
        
        # 添加各个功能页
        self.tab_widget.addTab(self._create_keyword_page(), "📝 关键词回复")
        self.tab_widget.addTab(self._create_user_page(), "👥 用户管理")
        self.tab_widget.addTab(self._create_card_key_page(), "🎫 卡密管理")
        self.tab_widget.addTab(self._create_log_page(), "📋 日志")
        
        layout.addWidget(self.tab_widget)
        
        # 保存UI引用
        self._ui_widget = widget
        
        return widget
    
    def _create_keyword_page(self):
        """创建关键词回复页面"""
        page = QWidget()
        layout = QVBoxLayout()
        page.setLayout(layout)
        
        # 添加关键词区域
        add_group = QGroupBox("添加关键词")
        add_layout = QVBoxLayout()
        add_group.setLayout(add_layout)
        
        # 关键词输入
        keyword_layout = QHBoxLayout()
        keyword_layout.addWidget(QLabel("关键词:"))
        self.keyword_input = QLineEdit()
        self.keyword_input.setPlaceholderText("输入关键词（例如：你好）")
        keyword_layout.addWidget(self.keyword_input)
        add_layout.addLayout(keyword_layout)
        
        # 回复话术输入
        reply_layout = QVBoxLayout()
        reply_layout.addWidget(QLabel("回复话术:"))
        self.reply_input = QTextEdit()
        self.reply_input.setPlaceholderText("输入回复话术（例如：您好，有什么可以帮助您的？）")
        self.reply_input.setMaximumHeight(100)
        reply_layout.addWidget(self.reply_input)
        add_layout.addLayout(reply_layout)
        
        # 添加按钮
        add_btn = QPushButton("➕ 添加关键词")
        add_btn.setStyleSheet("background-color: #52c41a; color: white; font-weight: bold; padding: 8px;")
        add_btn.clicked.connect(self._on_add_keyword)
        add_layout.addWidget(add_btn)
        
        layout.addWidget(add_group)
        
        # 关键词列表
        list_group = QGroupBox("关键词列表")
        list_layout = QVBoxLayout()
        list_group.setLayout(list_layout)
        
        self.keyword_table = QTableWidget()
        self.keyword_table.setColumnCount(3)
        self.keyword_table.setHorizontalHeaderLabels(["关键词", "回复话术", "操作"])
        
        header = self.keyword_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        
        list_layout.addWidget(self.keyword_table)
        
        layout.addWidget(list_group)
        
        # 加载数据
        self._update_keyword_table()
        
        return page
    
    def _create_user_page(self):
        """创建用户管理页面"""
        page = QWidget()
        layout = QVBoxLayout()
        page.setLayout(layout)
        
        # 说明区域
        info_group = QGroupBox("📌 开户说明")
        info_layout = QVBoxLayout()
        info_group.setLayout(info_layout)
        
        info_label = QLabel(
            "用户发送 <b>\"聚合开户\"</b> 命令即可自动开户（精准匹配）\n"
            "• 开户赠送 10.00 元余额\n"
            "• 仅支持私聊消息\n"
            "• 每个用户只能开户一次"
        )
        info_label.setStyleSheet("padding: 10px; color: #555;")
        info_layout.addWidget(info_label)
        
        layout.addWidget(info_group)
        
        # 用户列表
        list_group = QGroupBox("用户列表")
        list_layout = QVBoxLayout()
        list_group.setLayout(list_layout)
        
        self.user_table = QTableWidget()
        self.user_table.setColumnCount(7)
        self.user_table.setHorizontalHeaderLabels(["昵称", "WXID", "开户时间", "余额", "用户状态", "会员身份", "操作"])
        
        # 启用右键菜单
        self.user_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.user_table.customContextMenuRequested.connect(self._show_user_context_menu)
        
        header = self.user_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        
        list_layout.addWidget(self.user_table)
        
        layout.addWidget(list_group)
        
        # 加载数据
        self._update_user_table()
        
        return page
    
    def _create_card_key_page(self):
        """创建卡密管理页面"""
        page = QWidget()
        layout = QVBoxLayout()
        page.setLayout(layout)
        
        # 说明区域
        info_group = QGroupBox("📌 卡密管理说明")
        info_layout = QVBoxLayout()
        info_group.setLayout(info_layout)
        
        info_label = QLabel(
            "• 卡密格式：JHKM + 16位MD5\n"
            "• 用户发送卡密即可自动充值到余额\n"
            "• 管理员可通过发送 <b>\"生成卡密@金额\"</b> 命令生成卡密\n"
            "• 示例：生成卡密@30 （生成30￥面额的卡密）"
        )
        info_label.setStyleSheet("padding: 10px; color: #555;")
        info_layout.addWidget(info_label)
        
        layout.addWidget(info_group)
        
        # 生成卡密区域
        generate_group = QGroupBox("🎫 批量生成卡密")
        generate_layout = QHBoxLayout()
        generate_group.setLayout(generate_layout)
        
        # 面额输入
        generate_layout.addWidget(QLabel("面额（￥）："))
        self.card_amount_input = QDoubleSpinBox()
        self.card_amount_input.setMinimum(0.01)
        self.card_amount_input.setMaximum(10000)
        self.card_amount_input.setValue(10)
        self.card_amount_input.setFont(QFont("微软雅黑", 11))
        self.card_amount_input.setMinimumHeight(35)
        generate_layout.addWidget(self.card_amount_input)
        
        # 数量输入
        generate_layout.addWidget(QLabel("数量："))
        self.card_count_input = QSpinBox()
        self.card_count_input.setMinimum(1)
        self.card_count_input.setMaximum(1000)
        self.card_count_input.setValue(1)
        self.card_count_input.setFont(QFont("微软雅黑", 11))
        self.card_count_input.setMinimumHeight(35)
        generate_layout.addWidget(self.card_count_input)
        
        # 生成按钮
        generate_btn = QPushButton("🎁 批量生成")
        generate_btn.setFont(QFont("微软雅黑", 11))
        generate_btn.setMinimumHeight(40)
        generate_btn.setStyleSheet("""
            QPushButton {
                background-color: #52c41a;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 4px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #73d13d;
            }
            QPushButton:pressed {
                background-color: #389e0d;
            }
        """)
        generate_btn.clicked.connect(self._generate_card_keys_batch)
        generate_layout.addWidget(generate_btn)
        
        generate_layout.addStretch()
        
        layout.addWidget(generate_group)
        
        # 筛选和刷新区域
        filter_layout = QHBoxLayout()
        
        filter_layout.addWidget(QLabel("状态筛选："))
        self.card_filter_combo = QComboBox()
        self.card_filter_combo.addItems(["全部", "未使用", "已使用"])
        self.card_filter_combo.setFont(QFont("微软雅黑", 11))
        self.card_filter_combo.setMinimumHeight(35)
        self.card_filter_combo.currentIndexChanged.connect(self._filter_card_keys)
        filter_layout.addWidget(self.card_filter_combo)
        
        # 刷新按钮
        refresh_btn = QPushButton("🔄 刷新")
        refresh_btn.setFont(QFont("微软雅黑", 11))
        refresh_btn.setMinimumHeight(40)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #1890ff;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 4px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #40a9ff;
            }
            QPushButton:pressed {
                background-color: #096dd9;
            }
        """)
        refresh_btn.clicked.connect(self._update_card_key_table)
        filter_layout.addWidget(refresh_btn)
        
        filter_layout.addStretch()
        
        layout.addLayout(filter_layout)
        
        # 卡密列表
        list_group = QGroupBox("卡密列表")
        list_layout = QVBoxLayout()
        list_group.setLayout(list_layout)
        
        self.card_key_table = QTableWidget()
        self.card_key_table.setColumnCount(6)
        self.card_key_table.setHorizontalHeaderLabels(["卡密", "面额", "创建时间", "状态", "使用时间", "使用者"])
        
        # 设置表格样式和字体
        self.card_key_table.setFont(QFont("微软雅黑", 10))
        self.card_key_table.verticalHeader().setDefaultSectionSize(40)
        self.card_key_table.setAlternatingRowColors(True)
        self.card_key_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #e0e0e0;
                font-size: 11px;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 8px;
                font-weight: bold;
                font-size: 12px;
                border: 1px solid #d0d0d0;
            }
        """)
        
        header = self.card_key_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        
        list_layout.addWidget(self.card_key_table)
        
        layout.addWidget(list_group)
        
        # 加载数据
        self._update_card_key_table()
        
        return page
    
    def _generate_card_keys_batch(self):
        """批量生成卡密（UI按钮点击）"""
        try:
            amount = self.card_amount_input.value()
            count = self.card_count_input.value()
            
            # 生成卡密
            generated_keys = self._generate_card_key(amount, count)
            
            # 记录日志
            self.add_log("INFO", f"批量生成卡密 {count} 个，面额 {amount} ￥")
            
            # 刷新表格
            self._update_card_key_table()
            
            # 显示成功消息
            QMessageBox.information(
                self._ui_widget,
                "生成成功",
                f"成功生成 {count} 个面额为 {amount} ￥的卡密！\n\n卡密列表：\n" + "\n".join(generated_keys)
            )
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"批量生成卡密失败: {e}\n{traceback.format_exc()}")
            QMessageBox.critical(self._ui_widget, "生成失败", f"生成卡密时出错：{str(e)}")
    
    def _filter_card_keys(self):
        """筛选卡密"""
        self._update_card_key_table()
    
    def _update_card_key_table(self):
        """更新卡密表格"""
        self.card_key_table.setRowCount(0)
        
        # 获取筛选条件
        filter_text = self.card_filter_combo.currentText()
        
        for card_info in self.card_keys:
            # 应用筛选
            if filter_text == "未使用" and card_info["used"]:
                continue
            elif filter_text == "已使用" and not card_info["used"]:
                continue
            
            row = self.card_key_table.rowCount()
            self.card_key_table.insertRow(row)
            
            # 卡密
            self.card_key_table.setItem(row, 0, QTableWidgetItem(card_info["card_key"]))
            
            # 面额
            amount_item = QTableWidgetItem(f"{card_info['amount']:.2f} ￥")
            amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.card_key_table.setItem(row, 1, amount_item)
            
            # 创建时间
            self.card_key_table.setItem(row, 2, QTableWidgetItem(card_info["create_time"]))
            
            # 状态
            status_item = QTableWidgetItem("已使用" if card_info["used"] else "未使用")
            if card_info["used"]:
                status_item.setForeground(Qt.gray)
            else:
                status_item.setForeground(Qt.darkGreen)
            self.card_key_table.setItem(row, 3, status_item)
            
            # 使用时间
            used_time = card_info.get("used_time") or "-"
            self.card_key_table.setItem(row, 4, QTableWidgetItem(used_time))
            
            # 使用者
            used_by = card_info.get("used_by") or "-"
            self.card_key_table.setItem(row, 5, QTableWidgetItem(used_by))
    
    def _create_log_page(self):
        """创建日志页面"""
        page = QWidget()
        layout = QVBoxLayout()
        page.setLayout(layout)
        
        # 操作按钮
        button_layout = QHBoxLayout()
        
        clear_btn = QPushButton("🗑️ 清空日志")
        clear_btn.setStyleSheet("background-color: #ff4d4f; color: white; font-weight: bold; padding: 8px;")
        clear_btn.clicked.connect(self._on_clear_logs)
        button_layout.addWidget(clear_btn)
        
        export_btn = QPushButton("📤 导出日志")
        export_btn.setStyleSheet("background-color: #1890ff; color: white; font-weight: bold; padding: 8px;")
        export_btn.clicked.connect(self._on_export_logs)
        button_layout.addWidget(export_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        # 日志列表
        log_group = QGroupBox("日志记录")
        log_layout = QVBoxLayout()
        log_group.setLayout(log_layout)
        
        self.log_table = QTableWidget()
        self.log_table.setColumnCount(3)
        self.log_table.setHorizontalHeaderLabels(["时间", "级别", "消息"])
        
        header = self.log_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        
        log_layout.addWidget(self.log_table)
        
        layout.addWidget(log_group)
        
        # 加载数据
        self._update_log_table()
        
        return page
    
    # ==================== 关键词回复相关方法 ====================
    
    def _on_add_keyword(self):
        """添加关键词"""
        keyword = self.keyword_input.text().strip()
        reply = self.reply_input.toPlainText().strip()
        
        if not keyword:
            QMessageBox.warning(None, "提示", "请输入关键词")
            return
        
        if not reply:
            QMessageBox.warning(None, "提示", "请输入回复话术")
            return
        
        # 检查关键词是否已存在
        for item in self.keywords:
            if item["keyword"] == keyword:
                QMessageBox.warning(None, "提示", f"关键词 '{keyword}' 已存在")
                return
        
        # 添加到列表
        self.keywords.append({
            "keyword": keyword,
            "reply": reply
        })
        
        # 立即保存到配置，防止数据丢失
        self.set_config("keywords", self.keywords)
        
        # 清空输入
        self.keyword_input.clear()
        self.reply_input.clear()
        
        # 更新表格
        self._update_keyword_table()
        
        # 记录日志
        self.add_log("INFO", f"添加关键词: {keyword}")
        
        QMessageBox.information(None, "成功", "关键词添加成功")
    
    def _on_delete_keyword(self, row):
        """删除关键词"""
        if row < 0 or row >= len(self.keywords):
            return
        
        keyword_data = self.keywords[row]
        reply = QMessageBox.question(
            None,
            "确认删除",
            f"确定要删除关键词 '{keyword_data['keyword']}' 吗？",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            del self.keywords[row]
            
            # 立即保存到配置，防止数据丢失
            self.set_config("keywords", self.keywords)
            
            self._update_keyword_table()
            self.add_log("INFO", f"删除关键词: {keyword_data['keyword']}")
    
    def _update_keyword_table(self):
        """更新关键词表格"""
        self.keyword_table.setRowCount(0)
        
        for idx, item in enumerate(self.keywords):
            row = self.keyword_table.rowCount()
            self.keyword_table.insertRow(row)
            
            # 关键词
            self.keyword_table.setItem(row, 0, QTableWidgetItem(item["keyword"]))
            
            # 回复话术
            self.keyword_table.setItem(row, 1, QTableWidgetItem(item["reply"]))
            
            # 操作按钮
            delete_btn = QPushButton("删除")
            delete_btn.setStyleSheet("background-color: #ff4d4f; color: white;")
            delete_btn.clicked.connect(lambda checked, r=idx: self._on_delete_keyword(r))
            self.keyword_table.setCellWidget(row, 2, delete_btn)
    
    # ==================== 用户管理UI相关方法 ====================
    
    def _show_user_context_menu(self, pos):
        """显示用户右键菜单"""
        # 获取点击的行
        row = self.user_table.rowAt(pos.y())
        if row < 0 or row >= len(self.users):
            return
        
        user = self.users[row]
        
        # 创建右键菜单
        menu = QMenu()
        
        # 封禁/解封用户
        if user["status"] == "正常":
            ban_action = QAction("🚫 封禁用户", None)
            ban_action.triggered.connect(lambda: self._on_ban_user_ui(user["wxid"]))
            menu.addAction(ban_action)
        else:
            unban_action = QAction("✅ 解封用户", None)
            unban_action.triggered.connect(lambda: self._on_unban_user_ui(user["wxid"]))
            menu.addAction(unban_action)
        
        menu.addSeparator()
        
        # 添加余额
        add_balance_action = QAction("💰 添加余额", None)
        add_balance_action.triggered.connect(lambda: self._on_add_balance_ui(user["wxid"]))
        menu.addAction(add_balance_action)
        
        # 扣除余额
        deduct_balance_action = QAction("💸 扣除余额", None)
        deduct_balance_action.triggered.connect(lambda: self._on_deduct_balance_ui(user["wxid"]))
        menu.addAction(deduct_balance_action)
        
        menu.addSeparator()
        
        # 删除用户
        delete_action = QAction("🗑️ 删除用户", None)
        delete_action.triggered.connect(lambda: self._on_delete_user(row))
        menu.addAction(delete_action)
        
        # 显示菜单
        menu.exec_(self.user_table.viewport().mapToGlobal(pos))
    
    def _on_ban_user_ui(self, wxid):
        """封禁用户（UI操作）"""
        reply = QMessageBox.question(
            None,
            "确认封禁",
            f"确定要封禁用户 {wxid} 吗？\n封禁后该用户将无法使用功能。",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            if self._ban_user(wxid):
                self.add_log("INFO", f"已封禁用户: {wxid}")
                QMessageBox.information(None, "成功", "用户已封禁")
            else:
                QMessageBox.warning(None, "失败", "封禁用户失败")
    
    def _on_unban_user_ui(self, wxid):
        """解封用户（UI操作）"""
        reply = QMessageBox.question(
            None,
            "确认解封",
            f"确定要解封用户 {wxid} 吗？",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            if self._unban_user(wxid):
                self.add_log("INFO", f"已解封用户: {wxid}")
                QMessageBox.information(None, "成功", "用户已解封")
            else:
                QMessageBox.warning(None, "失败", "解封用户失败")
    
    def _on_add_balance_ui(self, wxid):
        """添加余额（UI操作）"""
        amount, ok = QInputDialog.getDouble(
            None,
            "添加余额",
            f"请输入要添加的金额（元）：",
            0.0,  # 默认值
            0.0,  # 最小值
            999999.0,  # 最大值
            2  # 小数位数
        )
        
        if ok and amount > 0:
            if self._add_balance(wxid, amount):
                self.add_log("INFO", f"已为用户 {wxid} 添加余额 {amount:.2f}元")
                QMessageBox.information(None, "成功", f"成功添加 {amount:.2f} 元")
            else:
                QMessageBox.warning(None, "失败", "添加余额失败")
    
    def _on_deduct_balance_ui(self, wxid):
        """扣除余额（UI操作）"""
        # 获取用户当前余额
        user_balance = 0
        for user in self.users:
            if user["wxid"] == wxid:
                user_balance = user["balance"]
                break
        
        amount, ok = QInputDialog.getDouble(
            None,
            "扣除余额",
            f"当前余额: {user_balance:.2f} 元\n请输入要扣除的金额（元）：",
            0.0,  # 默认值
            0.0,  # 最小值
            user_balance,  # 最大值（不能超过当前余额）
            2  # 小数位数
        )
        
        if ok and amount > 0:
            if self._deduct_balance(wxid, amount):
                self.add_log("INFO", f"已扣除用户 {wxid} 余额 {amount:.2f}元")
                QMessageBox.information(None, "成功", f"成功扣除 {amount:.2f} 元")
            else:
                QMessageBox.warning(None, "失败", "扣除余额失败（余额不足）")
    
    def _on_delete_user(self, row):
        """删除用户"""
        if row < 0 or row >= len(self.users):
            return
        
        user_data = self.users[row]
        reply = QMessageBox.question(
            None,
            "确认删除",
            f"确定要删除用户 '{user_data['nickname']}' ({user_data['wxid']}) 吗？",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            del self.users[row]
            self._update_user_table()
            self.add_log("INFO", f"删除用户: {user_data['nickname']} ({user_data['wxid']})")
    
    def _safe_update_user_table(self):
        """安全更新用户表格（线程安全）"""
        try:
            if self._ui_widget and hasattr(self, 'user_table'):
                QTimer.singleShot(0, self._update_user_table)
            else:
                # UI还未创建，稍后会自动加载
                pass
        except Exception as e:
            self.add_log("ERROR", f"调度UI更新失败: {e}")
    
    def _update_user_table(self):
        """更新用户表格"""
        self.user_table.setRowCount(0)
        
        for idx, user in enumerate(self.users):
            row = self.user_table.rowCount()
            self.user_table.insertRow(row)
            
            # 昵称
            self.user_table.setItem(row, 0, QTableWidgetItem(user["nickname"]))
            
            # WXID
            wxid = user["wxid"]
            self.user_table.setItem(row, 1, QTableWidgetItem(wxid))
            
            # 开户时间
            self.user_table.setItem(row, 2, QTableWidgetItem(user["open_time"]))
            
            # 余额
            balance_item = QTableWidgetItem(f"{user['balance']:.2f} 元")
            balance_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.user_table.setItem(row, 3, balance_item)
            
            # 用户状态
            status_item = QTableWidgetItem(user["status"])
            # 根据状态设置颜色
            if user["status"] == "正常":
                status_item.setForeground(Qt.darkGreen)
            elif user["status"] == "禁用":
                status_item.setForeground(Qt.red)
            elif user["status"] == "冻结":
                status_item.setForeground(Qt.blue)
            else:
                status_item.setForeground(Qt.gray)
            self.user_table.setItem(row, 4, status_item)
            
            # 会员身份
            is_vip, vip_type, expire_time = self._check_vip_status(wxid)
            if is_vip:
                vip_text = f"{vip_type}\n到期:{expire_time}"
                vip_item = QTableWidgetItem(vip_text)
                # 根据会员类型设置颜色
                if vip_type == "年卡":
                    vip_item.setForeground(Qt.darkRed)  # 深红色
                elif vip_type == "季卡":
                    vip_item.setForeground(Qt.darkMagenta)  # 深紫色
                elif vip_type == "月卡":
                    vip_item.setForeground(Qt.darkCyan)  # 深青色
            else:
                vip_item = QTableWidgetItem("普通用户")
                vip_item.setForeground(Qt.gray)
            self.user_table.setItem(row, 5, vip_item)
            
            # 操作按钮
            delete_btn = QPushButton("删除")
            delete_btn.setStyleSheet("background-color: #ff4d4f; color: white;")
            delete_btn.clicked.connect(lambda checked, r=idx: self._on_delete_user(r))
            self.user_table.setCellWidget(row, 6, delete_btn)
    
    # ==================== 日志相关方法 ====================
    
    def _on_clear_logs(self):
        """清空日志"""
        reply = QMessageBox.question(
            None,
            "确认清空",
            "确定要清空所有日志吗？",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.logs.clear()
            
            # 重置日志保存计数器
            self.log_save_counter = 0
            
            # 立即保存到文件，防止数据丢失
            self._save_logs_to_file()
            
            self._update_log_table()
            QMessageBox.information(None, "成功", "日志已清空")
    
    def _on_export_logs(self):
        """导出日志"""
        # 这里预留导出日志的功能
        # 后续可以实现导出到文件
        QMessageBox.information(None, "提示", "导出日志功能待实现")
    
    def _update_log_table(self):
        """更新日志表格"""
        self.log_table.setRowCount(0)
        
        # 倒序显示（最新的在上面）
        for log_entry in reversed(self.logs[-100:]):  # 只显示最近100条
            row = self.log_table.rowCount()
            self.log_table.insertRow(row)
            
            # 时间
            self.log_table.setItem(row, 0, QTableWidgetItem(log_entry["time"]))
            
            # 级别
            level_item = QTableWidgetItem(log_entry["level"])
            if log_entry["level"] == "ERROR":
                level_item.setForeground(Qt.red)
            elif log_entry["level"] == "WARNING":
                level_item.setForeground(Qt.darkYellow)
            else:
                level_item.setForeground(Qt.darkGreen)
            self.log_table.setItem(row, 1, level_item)
            
            # 消息
            self.log_table.setItem(row, 2, QTableWidgetItem(log_entry["message"]))
    
    # ==================== 用户管理相关方法 ====================
    
    def _handle_check_balance(self, instance_wxid, from_wxid):
        """处理用户查询余额"""
        try:
            # 查找用户
            user = None
            for u in self.users:
                if u["wxid"] == from_wxid:
                    user = u
                    break
            
            if not user:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 您还未开户，请先发送\"聚合开户\"进行注册\n\n发送\"教程\"查看使用教程")
                return
            
            balance = user.get("balance", 0.0)
            status = user.get("status", "正常")
            
            msg = f"💰 您的账户\n\n当前余额: {balance:.2f}￥"
            self._safe_send_text(instance_wxid, from_wxid, msg)
            
            self.add_log("INFO", f"用户 {from_wxid} 查询余额: {balance:.2f}￥")
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理余额查询时出错: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, "❌ 查询余额失败，请稍后重试")
    
    def _handle_register_user(self, instance_wxid, from_wxid):
        """处理用户开户"""
        try:
            self.add_log("INFO", f"开始处理聚合开户: {from_wxid}")
            
            # 检查用户是否已开户
            for user in self.users:
                if user["wxid"] == from_wxid:
                    self.add_log("INFO", f"用户已聚合开户: {from_wxid}")
                    self._safe_send_text(instance_wxid, from_wxid, f"您已经聚合开户了！\n当前余额: {user['balance']:.2f}元\n状态: {user['status']}")
                    return
            
            # 创建新用户
            new_user = {
                "nickname": from_wxid,  # 暂时使用wxid作为昵称
                "wxid": from_wxid,
                "open_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "balance": 10.0,  # 赠送10元
                "status": "正常"
            }
            
            self.add_log("INFO", f"创建新用户记录: {from_wxid}")
            self.users.append(new_user)
            self.set_config("users", self.users)
            
            self.add_log("INFO", f"新用户聚合开户成功: {from_wxid}")
            self.add_log("INFO", f"当前用户总数: {len(self.users)}")
            
            # 使用QTimer在主线程中更新UI（避免跨线程UI操作崩溃）
            self._safe_update_user_table()
            
            # 发送开户成功消息
            msg = f"🎉 聚合开户成功！\n\n欢迎使用本服务！\n赠送余额: 10.00元\n状态: 正常\n\n您现在可以使用各项功能了！"
            self._safe_send_text(instance_wxid, from_wxid, msg)
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            self.add_log("ERROR", f"处理聚合开户时出错: {e}\n{error_detail}")
            try:
                self._safe_send_text(instance_wxid, from_wxid, "聚合开户失败，请稍后重试")
            except:
                pass  # 避免二次崩溃
    
    def _handle_gm_coupon(self, instance_wxid, from_wxid, content):
        """处理商家券码撤销消息"""
        try:
            if not GM_AVAILABLE or not generate_coupon_image:
                self.add_log("ERROR", "缺少GenerateCoupon模块，无法生成券码图片")
                return
            
            self.add_log("INFO", f"收到券码撤销消息，来自: {from_wxid}")
            
            # 解析消息内容，提取标题和券码
            # 格式：【美团】商家撤销了您的券码，您的「熊喵来了火锅代金券」券码已恢复，新券码为024480932477。
            title_match = re.search(r'「(.+?)」', content)
            code_match = re.search(r'新券码为(\d+)', content)
            
            # 券码是必需的，如果无法提取则返回
            if not code_match:
                self.add_log("ERROR", f"无法从消息中提取券码: {content[:100]}")
                return
            
            coupon_code = code_match.group(1)  # 例如：024480932477
            
            # 标题如果无法提取，使用默认值"团购券"
            if title_match:
                title = title_match.group(1)  # 例如：熊喵来了火锅代金券
            else:
                title = "团购券"
                self.add_log("WARNING", f"无法从消息中提取标题，使用默认标题: {title}")
            
            self.add_log("INFO", f"提取到标题: {title}, 券码: {coupon_code}")
            
            # 格式化券码（带空格分隔）
            formatted_code = format_voucher_code(coupon_code) if format_voucher_code else coupon_code
            
            # 生成券码图片（使用默认有效期）
            # main(title, barcode_text, code, notes, date_text)
            save_path = generate_coupon_image(title, coupon_code, formatted_code, "", "")
            
            if save_path and os.path.exists(save_path):
                self.add_log("INFO", f"券码图片已生成: {save_path}")
                
                # 添加延迟
                time.sleep(0.5)
                
                # 发送图片
                result = self.send_image(instance_wxid, from_wxid, save_path)
                
                if result.get("success"):
                    self.add_log("INFO", f"券码图片发送成功")
                else:
                    error = result.get("error", "未知错误")
                    self.add_log("ERROR", f"券码图片发送失败: {error}")
            else:
                self.add_log("ERROR", "券码图片生成失败或文件不存在")
                
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理GM券码时出错: {e}\n{traceback.format_exc()}")
    
    def _handle_manual_coupon(self, instance_wxid, from_wxid, content):
        """处理作图命令"""
        try:
            if not GM_AVAILABLE or not generate_coupon_image:
                self._safe_send_text(instance_wxid, from_wxid, "系统错误：缺少GenerateCoupon模块")
                return
            
            # 移除"作图"前缀
            content = content[2:].strip()
            
            if not content:
                self._safe_send_text(instance_wxid, from_wxid, "格式错误！\n正确格式：\n作图标题\n券码\n\n或者直接发送券码（标题默认为\"团购券\"）：\n作图\n券码\n\n示例1：\n作图龙歌团购券\n010123456789\n\n示例2（无标题）：\n作图\n010123456789\n\n如需帮助请发送\"教程\"查看使用教程")
                return
            
            # 按换行符分割
            lines = [line.strip() for line in content.split('\n') if line.strip()]
            
            if len(lines) == 0:
                self._safe_send_text(instance_wxid, from_wxid, "券码不能为空！")
                return
            
            # 判断第一行是标题还是券码
            # 券码特征：纯数字或数字字母组合，长度>=10，不含中文
            # 标题特征：包含中文或其他字符
            first_line = lines[0]
            
            # 检查是否包含中文字符
            def contains_chinese(text):
                for char in text:
                    if '\u4e00' <= char <= '\u9fff':
                        return True
                return False
            
            # 判断是否为券码
            is_coupon = False
            if not contains_chinese(first_line):
                # 不含中文，检查是否像券码（纯数字或数字字母组合，长度>=10）
                if first_line.isdigit() and len(first_line) >= 10:
                    is_coupon = True
                elif first_line.isalnum() and len(first_line) >= 10:
                    # 检查是否主要是数字
                    digit_count = sum(c.isdigit() for c in first_line)
                    if digit_count >= len(first_line) * 0.7:  # 至少70%是数字
                        is_coupon = True
            
            if is_coupon:
                # 第一行是券码，标题默认为"团购券"
                title = "团购券"
                coupon_lines = lines
            else:
                # 第一行是标题
                title = first_line
                coupon_lines = lines[1:]
            
            # 检查是否有券码
            if len(coupon_lines) == 0:
                self._safe_send_text(instance_wxid, from_wxid, "券码不能为空！")
                return
            
            # 检查批量作图数量限制
            if len(coupon_lines) > 5:
                self.add_log("INFO", f"用户 {from_wxid} 批量作图超过限制: {len(coupon_lines)}张")
                self._safe_send_text(instance_wxid, from_wxid, "抱歉，批量作图限制最多五张")
                return
            
            # 判断是单张还是批量
            is_batch = len(coupon_lines) > 1
            
            if is_batch:
                self.add_log("INFO", f"开始批量生成券码图片，标题: {title}, 数量: {len(coupon_lines)}")
                self._safe_send_text(instance_wxid, from_wxid, f"正在批量生成{len(coupon_lines)}张图片，请稍候...")
            else:
                self.add_log("INFO", f"开始生成券码图片，标题: {title}, 券码: {coupon_lines[0]}")
                self._safe_send_text(instance_wxid, from_wxid, "正在生成图片，请稍候...")
            
            # 在子线程中生成图片并发送
            def generate_and_send():
                try:
                    success_count = 0
                    fail_count = 0
                    
                    for idx, coupon_code in enumerate(coupon_lines):
                        try:
                            # 格式化券码
                            formatted_code = format_voucher_code(coupon_code) if format_voucher_code else coupon_code
                            
                            # 生成图片（不显示有效期）
                            # main(title, barcode_text, code, notes, date_text)
                            image_path = generate_coupon_image(title, coupon_code, formatted_code, "", "")
                            
                            if image_path and os.path.exists(image_path):
                                self.add_log("INFO", f"券码图片生成成功 ({idx+1}/{len(coupon_lines)}): {image_path}")
                                
                                # 发送图片
                                time.sleep(0.3)  # 批量发送时减少延迟
                                result = self.send_image(instance_wxid, from_wxid, image_path)
                                
                                if result.get("success"):
                                    self.add_log("INFO", f"券码图片发送成功 ({idx+1}/{len(coupon_lines)})")
                                    success_count += 1
                                else:
                                    error = result.get("error", "未知错误")
                                    self.add_log("ERROR", f"券码图片发送失败 ({idx+1}/{len(coupon_lines)}): {error}")
                                    fail_count += 1
                            else:
                                self.add_log("ERROR", f"券码图片生成失败 ({idx+1}/{len(coupon_lines)})")
                                fail_count += 1
                            
                            # 批量发送时，每张图片之间添加短暂延迟
                            if is_batch and idx < len(coupon_lines) - 1:
                                time.sleep(0.5)
                        
                        except Exception as e:
                            self.add_log("ERROR", f"处理券码 {coupon_code} 时出错: {e}")
                            fail_count += 1
                    
                    # 发送完成提示
                    if is_batch:
                        if fail_count == 0:
                            self._safe_send_text(instance_wxid, from_wxid, f"✅ 批量作图完成！成功生成{success_count}张图片")
                        else:
                            self._safe_send_text(instance_wxid, from_wxid, f"批量作图完成！成功{success_count}张，失败{fail_count}张")
                    else:
                        if fail_count > 0:
                            self._safe_send_text(instance_wxid, from_wxid, "图片生成失败，请稍后重试")
                        
                except Exception as e:
                    import traceback
                    self.add_log("ERROR", f"生成券码图片时出错: {e}\n{traceback.format_exc()}")
                    self._safe_send_text(instance_wxid, from_wxid, f"生成图片失败: {e}")
            
            # 启动子线程
            thread = threading.Thread(target=generate_and_send)
            thread.daemon = True
            thread.start()
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理作图命令时出错: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
    
    def _handle_goofish_kami(self, instance_wxid, from_wxid, content):
        """处理闲管家卡密链接"""
        try:
            if not GM_AVAILABLE or not generate_coupon_image:
                self._safe_send_text(instance_wxid, from_wxid, "系统错误：缺少GenerateCoupon模块")
                return
            
            # 提取token
            token_match = re.search(r'kami/([A-Za-z0-9_-]+)', content)
            if not token_match:
                self.add_log("ERROR", f"无法从链接中提取token: {content}")
                self._safe_send_text(instance_wxid, from_wxid, "无法识别卡密链接，请检查链接格式\n\n如需帮助请发送\"教程\"查看使用教程")
                return
            
            token = token_match.group(1)
            self.add_log("INFO", f"提取到token: {token}")
            self._safe_send_text(instance_wxid, from_wxid, "正在获取卡密信息，请稍候...")
            
            # 在子线程中请求API并生成图片
            def fetch_and_generate():
                try:
                    # 请求API
                    url = "https://api.goofish.pro/v1/h5/kami/order/detail"
                    headers = {
                        'accept': '*/*',
                        'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
                        'content-type': 'application/json',
                        'origin': 'https://m.goofish.pro',
                        'referer': 'https://m.goofish.pro/',
                        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36 Edg/140.0.0.0'
                    }
                    payload = {"token": token}
                    
                    self.add_log("INFO", f"正在请求API: {url}")
                    response = requests.post(url, json=payload, headers=headers, timeout=10, verify=False)
                    data = response.json()
                    
                    self.add_log("INFO", f"API响应: {json.dumps(data, ensure_ascii=False)[:200]}")
                    
                    if data.get("code") != 0:
                        error_msg = data.get("msg", "未知错误")
                        self.add_log("ERROR", f"API返回错误: {error_msg}")
                        self._safe_send_text(instance_wxid, from_wxid, f"获取卡密失败: {error_msg}")
                        return
                    
                    kind_list = data.get("data", {}).get("kind_list", [])
                    if not kind_list:
                        self.add_log("ERROR", "未找到卡密信息")
                        self._safe_send_text(instance_wxid, from_wxid, "未找到卡密信息")
                        return
                    
                    # 统计总券码数量
                    total_cards = sum(len(kind.get("card_list", [])) for kind in kind_list)
                    self.add_log("INFO", f"找到 {len(kind_list)} 个种类，共 {total_cards} 张券码")
                    
                    # 遍历每个kind，组装文本消息
                    for kind in kind_list:
                        kind_name = kind.get("kind_name", "未知券")
                        card_list = kind.get("card_list", [])
                        
                        if not card_list:
                            continue
                        
                        self.add_log("INFO", f"处理券种: {kind_name}, 数量: {len(card_list)}")
                        
                        # 组装文本消息：第一行是kind_name，后面每行一个券码
                        message_lines = [kind_name]
                        
                        # 收集所有券码
                        for card in card_list:
                            card_pwd = card.get("card_pwd", "")
                            if card_pwd:
                                message_lines.append(card_pwd)
                        
                        # 合并成一条消息
                        message_text = "\n".join(message_lines)
                        
                        # 发送文本消息
                        self.add_log("INFO", f"发送券码文本，共 {len(card_list)} 张")
                        self._safe_send_text(instance_wxid, from_wxid, message_text)
                    
                    # 全部完成
                    self.add_log("INFO", f"卡密处理完成，共发送 {total_cards} 张券码")
                    
                except requests.RequestException as e:
                    import traceback
                    self.add_log("ERROR", f"请求API失败: {e}\n{traceback.format_exc()}")
                    self._safe_send_text(instance_wxid, from_wxid, f"网络请求失败: {e}")
                except Exception as e:
                    import traceback
                    self.add_log("ERROR", f"处理卡密时出错: {e}\n{traceback.format_exc()}")
                    self._safe_send_text(instance_wxid, from_wxid, f"处理失败: {e}")
            
            # 启动子线程
            thread = threading.Thread(target=fetch_and_generate)
            thread.daemon = True
            thread.start()
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理闲管家卡密时出错: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
    
    def _handle_card_key_recharge(self, instance_wxid, from_wxid, card_key):
        """处理用户发送卡密充值"""
        try:
            # 使用卡密
            success, message, amount = self._use_card_key(card_key, from_wxid)
            
            # 发送结果消息
            self._safe_send_text(instance_wxid, from_wxid, message)
            
            # 如果充值成功，更新UI表格（在主线程中）
            if success and hasattr(self, '_update_user_table'):
                QTimer.singleShot(0, self._update_user_table)
        
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理卡密充值失败: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, f"❌ 处理卡密时出错：{str(e)}")
    
    def _handle_admin_generate_card_key(self, instance_wxid, from_wxid, content):
        """处理管理员生成卡密命令"""
        try:
            # 解析命令：生成卡密@金额
            parts = content.split("@")
            if len(parts) != 2:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 命令格式错误\n\n正确格式：生成卡密@金额\n示例：生成卡密@30")
                return
            
            # 解析金额
            try:
                amount = float(parts[1])
                if amount <= 0:
                    self._safe_send_text(instance_wxid, from_wxid, "❌ 金额必须大于0")
                    return
            except ValueError:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 金额格式错误，请输入数字\n\n示例：生成卡密@30")
                return
            
            # 生成卡密
            generated_keys = self._generate_card_key(amount, 1)
            
            if generated_keys:
                card_key = generated_keys[0]
                
                # 发送卡密给管理员
                msg = f"✅ 卡密生成成功！\n\n卡密：{card_key}\n面额：{amount}￥\n\n请将此卡密发送给用户进行充值"
                self._safe_send_text(instance_wxid, from_wxid, msg)
                
                # 记录日志
                self.add_log("INFO", f"管理员 {from_wxid} 生成了 {amount} ￥卡密: {card_key}")
            else:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 卡密生成失败，请稍后重试")
        
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"管理员生成卡密失败: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, f"❌ 生成卡密时出错：{str(e)}")
    
    def _handle_alliance_ranking(self, instance_wxid, from_wxid):
        """处理联盟榜单命令"""
        try:
            self.add_log("INFO", f"用户 {from_wxid} 请求联盟榜单")
            
            # 保存用户会话信息
            self.user_sessions[from_wxid] = {
                "type": "alliance_wait_list_type",
                "instance_wxid": instance_wxid
            }
            
            # 发送榜单类型选择菜单
            menu_msg = (
                "请选择榜单：\n"
                "1、今日必推\n"
                "2、同城热销\n"
                "3、实时热销\n\n"
                "回复对应编号获取榜单，如：1\n"
                "输入\"q\"退出会话"
            )
            self._safe_send_text(instance_wxid, from_wxid, menu_msg)
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理联盟榜单命令时出错: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
    
    def _handle_alliance_list_type_choice(self, instance_wxid, from_wxid, choice):
        """处理榜单类型选择"""
        try:
            # 检查是否是退出命令
            if choice.strip().lower() == "q":
                self.add_log("INFO", f"用户 {from_wxid} 退出联盟榜单")
                if from_wxid in self.user_sessions:
                    del self.user_sessions[from_wxid]
                self._safe_send_text(instance_wxid, from_wxid, "✅ 已退出联盟榜单功能")
                return
            
            # 验证选择
            list_type_map = {
                "1": (2, "今日必推"),
                "2": (3, "同城热销"),
                "3": (5, "实时热销")
            }
            
            if choice not in list_type_map:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 无效选择，请输入 1、2 或 3\n输入\"q\"退出会话")
                return
            
            list_topi_id, list_type_name = list_type_map[choice]
            
            # 更新会话状态
            self.user_sessions[from_wxid]["type"] = "alliance_wait_biz_line"
            self.user_sessions[from_wxid]["list_topi_id"] = list_topi_id
            self.user_sessions[from_wxid]["list_type_name"] = list_type_name
            
            self.add_log("INFO", f"用户 {from_wxid} 选择榜单类型: {list_type_name}")
            
            # 发送榜单分类选择菜单
            menu_msg = (
                "请选择榜单类型：\n"
                "1、到店餐饮\n"
                "2、到店综合\n\n"
                "回复对应编号获取榜单，如：1\n"
                "输入\"q\"退出会话"
            )
            self._safe_send_text(instance_wxid, from_wxid, menu_msg)
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理榜单类型选择时出错: {e}\n{traceback.format_exc()}")
            if from_wxid in self.user_sessions:
                del self.user_sessions[from_wxid]
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
    
    def _handle_alliance_biz_line_choice(self, instance_wxid, from_wxid, choice):
        """处理榜单分类选择"""
        try:
            # 检查是否是退出命令
            if choice.strip().lower() == "q":
                self.add_log("INFO", f"用户 {from_wxid} 退出联盟榜单")
                if from_wxid in self.user_sessions:
                    del self.user_sessions[from_wxid]
                self._safe_send_text(instance_wxid, from_wxid, "✅ 已退出联盟榜单功能")
                return
            
            # 验证选择
            biz_line_map = {
                "1": (1, "到店餐饮"),
                "2": (2, "到店综合")
            }
            
            if choice not in biz_line_map:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 无效选择，请输入 1 或 2\n输入\"q\"退出会话")
                return
            
            biz_line, biz_line_name = biz_line_map[choice]
            
            # 更新会话状态
            self.user_sessions[from_wxid]["type"] = "alliance_wait_city"
            self.user_sessions[from_wxid]["biz_line"] = biz_line
            self.user_sessions[from_wxid]["biz_line_name"] = biz_line_name
            
            self.add_log("INFO", f"用户 {from_wxid} 选择榜单分类: {biz_line_name}")
            
            # 发送城市输入提示
            menu_msg = (
                "请输入您查看榜单的城市：\n"
                "（如： 北京市、郑州市等，单次只能发一个城市）\n"
                "输入\"q\"退出会话"
            )
            self._safe_send_text(instance_wxid, from_wxid, menu_msg)
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理榜单分类选择时出错: {e}\n{traceback.format_exc()}")
            if from_wxid in self.user_sessions:
                del self.user_sessions[from_wxid]
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
    
    def _handle_alliance_city_input(self, instance_wxid, from_wxid, city_name):
        """处理城市输入"""
        try:
            # 检查是否是退出命令
            if city_name.strip().lower() == "q":
                self.add_log("INFO", f"用户 {from_wxid} 退出联盟榜单")
                if from_wxid in self.user_sessions:
                    del self.user_sessions[from_wxid]
                self._safe_send_text(instance_wxid, from_wxid, "✅ 已退出联盟榜单功能")
                return
            
            # 获取城市ID
            city_id, full_city_name = self._get_city_id(city_name.strip())
            
            if not city_id:
                self.add_log("WARNING", f"用户 {from_wxid} 输入的城市未找到: {city_name}")
                self._safe_send_text(instance_wxid, from_wxid, f"❌ 未找到城市：{city_name}\n请检查城市名称后重新输入\n输入\"q\"退出会话")
                return
            
            # 获取会话信息
            session = self.user_sessions.get(from_wxid)
            if not session:
                self._safe_send_text(instance_wxid, from_wxid, "会话已过期，请重新发送\"联盟榜单\"命令")
                return
            
            list_topi_id = session.get("list_topi_id")
            list_type_name = session.get("list_type_name")
            biz_line = session.get("biz_line")
            biz_line_name = session.get("biz_line_name")
            
            self.add_log("INFO", f"用户 {from_wxid} 输入城市: {full_city_name} (ID: {city_id})")
            
            # 清除会话
            if from_wxid in self.user_sessions:
                del self.user_sessions[from_wxid]
            
            # 发送查询中提示
            self._safe_send_text(instance_wxid, from_wxid, f"🔍 正在查询【{list_type_name} - {biz_line_name} - {full_city_name}】榜单，请稍候...")
            
            # 在子线程中查询（避免阻塞）
            def query_ranking():
                try:
                    # 调用query_products接口
                    params = {
                        "platform": 2,  # 到店
                        "bizLine": biz_line,
                        "sortField": 2,  # 按销量排序
                        "ascDescOrder": 2,  # 降序
                        "listTopiId": list_topi_id,
                        "cityId": city_id,
                        "pageSize": 20,  # 获取20个商品
                        "pageNo": 1
                    }
                    
                    self.add_log("INFO", f"调用query_products接口，参数: {params}")
                    
                    result = query_products(
                        app_key=self.mt_app_key,
                        app_secret=self.mt_app_secret,
                        query_params=params
                    )
                    
                    self.add_log("INFO", f"query_products返回结果: code={result.get('code')}, message={result.get('message')}")
                    
                    # 检查返回结果
                    if result.get("code") != 0:
                        error_msg = result.get("message", "未知错误")
                        self.add_log("ERROR", f"查询榜单失败: {error_msg}")
                        self._safe_send_text(instance_wxid, from_wxid, f"❌ 查询失败: {error_msg}")
                        return
                    
                    data = result.get("data", [])
                    
                    if not data:
                        self.add_log("INFO", f"查询榜单无数据")
                        self._safe_send_text(instance_wxid, from_wxid, "📭 该榜单暂无数据")
                        return
                    
                    self.add_log("INFO", f"查询到 {len(data)} 个商品")
                    
                    # 整理数据并发送
                    self._format_and_send_ranking_data(instance_wxid, from_wxid, data, list_type_name, biz_line_name, full_city_name)
                    
                except Exception as e:
                    import traceback
                    self.add_log("ERROR", f"查询榜单时出错: {e}\n{traceback.format_exc()}")
                    self._safe_send_text(instance_wxid, from_wxid, f"查询失败: {str(e)}")
            
            # 启动子线程
            thread = threading.Thread(target=query_ranking)
            thread.daemon = True
            thread.start()
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理城市输入时出错: {e}\n{traceback.format_exc()}")
            if from_wxid in self.user_sessions:
                del self.user_sessions[from_wxid]
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
    
    def _format_and_send_ranking_data(self, instance_wxid, from_wxid, data, list_type_name, biz_line_name, city_name):
        """整理榜单数据并发送"""
        try:
            # 获取会话信息中的biz_line（用于获取推广链接）
            biz_line = 1 if "餐饮" in biz_line_name else 2
            
            # 整理数据
            products = []
            for item in data:
                try:
                    brand_info = item.get("brandInfo", {})
                    coupon_detail = item.get("couponPackDetail", {})
                    commission_info = item.get("commissionInfo", {})
                    available_poi_info = item.get("availablePoiInfo", {})
                    coupon_valid_time_info = item.get("couponValidTimeInfo", {})
                    product_label = coupon_detail.get("productLabel", {})
                    price_power_label = product_label.get("pricePowerLabel", {}) if product_label else {}
                    
                    # 提取数据
                    brand_name = brand_info.get("brandName", "未知品牌")
                    coupon_name = coupon_detail.get("name", "未知卡券")
                    sell_price = coupon_detail.get("sellPrice", "0")
                    commission = commission_info.get("commission", "0")
                    sale_volume = coupon_detail.get("saleVolume", "0")
                    start_time = coupon_detail.get("startTime", 0)
                    end_time = coupon_detail.get("endTime", 0)
                    coupon_valid_day = coupon_valid_time_info.get("couponValidDay", 0)
                    sku_view_id = coupon_detail.get("skuViewId", "")
                    
                    # 额外字段
                    available_poi_city_num = available_poi_info.get("availablePoiCityNum", 0)
                    commission_percent = commission_info.get("commissionPercent", "0")
                    original_price = coupon_detail.get("originalPrice", "0")
                    history_price_label = price_power_label.get("historyPriceLabel", "") if price_power_label else ""
                    product_rank_label = product_label.get("productRankLabel", "") if product_label else ""
                    
                    # 时间戳转换
                    start_time_str = datetime.fromtimestamp(start_time).strftime("%Y-%m-%d %H:%M:%S") if start_time else ""
                    end_time_str = datetime.fromtimestamp(end_time).strftime("%Y-%m-%d %H:%M:%S") if end_time else ""
                    
                    # 获取推广短链
                    short_link = ""
                    if sku_view_id:
                        try:
                            link_params = {
                                "platform": 2,  # 到店
                                "bizLine": biz_line,
                                "skuViewId": sku_view_id,
                                "linkType": 2  # H5短链
                            }
                            link_result = get_referral_link(
                                app_key=self.mt_app_key,
                                app_secret=self.mt_app_secret,
                                request_data=link_params
                            )
                            if link_result.get("code") == 0:
                                short_link = link_result.get("data", "")
                            else:
                                self.add_log("WARNING", f"获取短链失败: {link_result.get('message')}")
                        except Exception as e:
                            self.add_log("WARNING", f"获取短链异常: {e}")
                    
                    products.append({
                        "brand_name": brand_name,
                        "coupon_name": coupon_name,
                        "sell_price": sell_price,
                        "commission": commission,
                        "sale_volume": sale_volume,
                        "start_time": start_time_str,
                        "end_time": end_time_str,
                        "coupon_valid_day": coupon_valid_day,
                        "available_poi_city_num": available_poi_city_num,
                        "commission_percent": commission_percent,
                        "original_price": original_price,
                        "history_price_label": history_price_label,
                        "product_rank_label": product_rank_label,
                        "short_link": short_link
                    })
                    
                except Exception as e:
                    self.add_log("WARNING", f"解析商品数据时出错: {e}")
                    continue
            
            if not products:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 数据解析失败")
                return
            
            # 只发送前10个商品的文字消息，分两批（每批5个）
            text_products = products[:10]  # 只取前10个用于文字消息
            for i in range(0, len(text_products), 5):
                batch = text_products[i:i+5]
                msg_lines = [f"📊 【{list_type_name} - {biz_line_name} - {city_name}】榜单（第{i//5+1}批）"]
                
                for idx, product in enumerate(batch, start=i+1):
                    msg_lines.append(f"\n🔹 商品 {idx}")
                    msg_lines.append(f"品牌名：{product['brand_name']}")
                    msg_lines.append(f"卡券名称：{product['coupon_name']}")
                    msg_lines.append(f"当前售价：{product['sell_price']}元")
                    msg_lines.append(f"佣金：{product['commission']}元")
                    msg_lines.append(f"实时销量：{product['sale_volume']}")
                    msg_lines.append(f"活动开始时间：{product['start_time']}")
                    msg_lines.append(f"活动结束时间：{product['end_time']}")
                    # 添加短链
                    if product['short_link']:
                        msg_lines.append(f"短链：{product['short_link']}")
                
                # 添加分隔线
                msg_lines.append("===================")
                
                msg = "\n".join(msg_lines)
                self._safe_send_text(instance_wxid, from_wxid, msg)
            
            # 生成Excel文件
            excel_path = self._generate_ranking_excel(products, list_type_name, biz_line_name, city_name)
            
            if excel_path:
                self.add_log("INFO", f"Excel文件生成成功: {excel_path}")
                
                # 发送Excel文件
                result = self.send_file(instance_wxid, from_wxid, excel_path)
                
                if result.get("success"):
                    self.add_log("INFO", f"Excel文件发送成功")
                    self._safe_send_text(instance_wxid, from_wxid, "✅ 榜单详细数据已发送")
                    
                    # 如果设置了发送后删除，则删除文件
                    if self.delete_after_send:
                        try:
                            os.remove(excel_path)
                            self.add_log("INFO", f"已删除Excel文件: {excel_path}")
                        except Exception as e:
                            self.add_log("WARNING", f"删除Excel文件失败: {e}")
                else:
                    self.add_log("ERROR", f"Excel文件发送失败: {result.get('message')}")
                    self._safe_send_text(instance_wxid, from_wxid, "⚠️ Excel文件发送失败")
            else:
                self._safe_send_text(instance_wxid, from_wxid, "⚠️ Excel文件生成失败，但文字信息已发送")
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"整理并发送榜单数据时出错: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
    
    def _generate_ranking_excel(self, products, list_type_name, biz_line_name, city_name):
        """生成榜单Excel文件"""
        try:
            # 检查openpyxl是否可用
            if not OPENPYXL_AVAILABLE:
                self.add_log("ERROR", "缺少openpyxl库，无法生成Excel文件")
                return None
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "联盟榜单"
            
            # 设置表头
            headers = [
                "序号",
                "品牌名",
                "卡券名称",
                "当前售价(元)",
                "佣金(元)",
                "佣金百分比",
                "原价(元)",
                "实时销量",
                "活动开始时间",
                "活动结束时间",
                "卡券有效时长(天)",
                "可用城市数量",
                "历史价格标签",
                "商品排名标签",
                "短链"
            ]
            
            ws.append(headers)
            
            # 填充数据
            for idx, product in enumerate(products, start=1):
                # 当couponValidDay为0时，Excel中显示空
                coupon_valid_day_display = "" if product["coupon_valid_day"] == 0 else product["coupon_valid_day"]
                
                row = [
                    idx,
                    product["brand_name"],
                    product["coupon_name"],
                    product["sell_price"],
                    product["commission"],
                    product["commission_percent"],
                    product["original_price"],
                    product["sale_volume"],
                    product["start_time"],
                    product["end_time"],
                    coupon_valid_day_display,
                    product["available_poi_city_num"],
                    product["history_price_label"],
                    product["product_rank_label"],
                    product["short_link"]
                ]
                ws.append(row)
            
            # 调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # 生成文件名
            timestamp = int(time.time())
            filename = f"联盟榜单-{list_type_name}-{biz_line_name}-{city_name}-{timestamp}.xlsx"
            
            # 保存到专门的Excel输出目录
            file_path = os.path.join(self.excel_output_dir, filename)
            
            # 确保目录存在
            os.makedirs(self.excel_output_dir, exist_ok=True)
            
            # 保存文件
            wb.save(file_path)
            
            self.add_log("INFO", f"联盟榜单Excel文件已保存到本地: {file_path}")
            
            return file_path
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"生成联盟榜单Excel文件时出错: {e}\n{traceback.format_exc()}")
            return None
    
    def _handle_91kami(self, instance_wxid, from_wxid, content):
        """处理91卡券链接"""
        try:
            # 提取token（从URL路径中提取，格式如：/cpd/token.aspx）
            token_match = re.search(r'mai\.91kami\.com/cpd/([A-Za-z0-9_-]+)\.aspx', content)
            if not token_match:
                self.add_log("ERROR", f"无法从链接中提取token: {content}")
                self._safe_send_text(instance_wxid, from_wxid, "无法识别91卡券链接，请检查链接格式\n\n如需帮助请发送\"教程\"查看使用教程")
                return
            
            token = token_match.group(1)
            self.add_log("INFO", f"提取到91卡券token: {token}")
            self._safe_send_text(instance_wxid, from_wxid, "正在获取卡券信息，请稍候...")
            
            # 在子线程中请求API并发送文本
            def fetch_and_send():
                try:
                    # 请求API
                    url = "https://mai.91kami.com/api/Cpd/Detail"
                    headers = {
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'zh-CN,zh;q=0.9',
                        'Connection': 'keep-alive',
                        'Content-Type': 'application/json',
                        'Origin': 'https://mai.91kami.com',
                        'Referer': 'https://mai.91kami.com/tq/',
                        'Sec-Fetch-Dest': 'empty',
                        'Sec-Fetch-Mode': 'cors',
                        'Sec-Fetch-Site': 'same-origin',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36',
                        'sec-ch-ua': '"Not;A=Brand";v="99", "Google Chrome";v="139", "Chromium";v="139"',
                        'sec-ch-ua-mobile': '?0',
                        'sec-ch-ua-platform': '"Windows"'
                    }
                    payload = {"token": token}
                    
                    self.add_log("INFO", f"正在请求91卡券API: {url}")
                    response = requests.post(url, json=payload, headers=headers, timeout=10, verify=False)
                    data = response.json()
                    
                    self.add_log("INFO", f"91卡券API响应: {json.dumps(data, ensure_ascii=False)[:200]}")
                    
                    # 检查响应数据
                    if "Data" not in data or not data["Data"]:
                        self.add_log("ERROR", "未找到卡券信息")
                        self._safe_send_text(instance_wxid, from_wxid, "未找到卡券信息")
                        return
                    
                    # Data可能是字符串，需要转换为列表
                    data_list = data["Data"]
                    if isinstance(data_list, str):
                        try:
                            data_list = json.loads(data_list)
                            self.add_log("INFO", f"91卡券Data已从字符串转换为列表，数量: {len(data_list) if isinstance(data_list, list) else 0}")
                        except json.JSONDecodeError as e:
                            self.add_log("ERROR", f"91卡券Data解析失败: {e}")
                            self._safe_send_text(instance_wxid, from_wxid, "卡券数据格式错误")
                            return
                    
                    # 确保data_list是列表
                    if not isinstance(data_list, list):
                        self.add_log("ERROR", f"91卡券Data不是列表类型: {type(data_list)}")
                        self._safe_send_text(instance_wxid, from_wxid, "卡券数据格式错误")
                        return
                    
                    # 统计总卡密数量
                    total_cards = sum(len(item.get("CardPwdArr", [])) for item in data_list)
                    self.add_log("INFO", f"找到 {len(data_list)} 个券种，共 {total_cards} 张卡密")
                    
                    # 遍历每个券种，组装文本消息
                    for item in data_list:
                        title = item.get("Title", "未知券")
                        card_pwd_arr = item.get("CardPwdArr", [])
                        
                        if not card_pwd_arr:
                            continue
                        
                        self.add_log("INFO", f"处理券种: {title}, 数量: {len(card_pwd_arr)}")
                        
                        # 组装文本消息：第一行是Title，后面每行一个卡密
                        message_lines = [title]
                        
                        # 收集所有卡密
                        for card in card_pwd_arr:
                            card_no = card.get("c", "")  # "c" 是卡密
                            if card_no:
                                message_lines.append(card_no)
                        
                        # 合并成一条消息
                        message_text = "\n".join(message_lines)
                        
                        # 发送文本消息
                        self.add_log("INFO", f"发送91卡券文本，共 {len(card_pwd_arr)} 张")
                        self._safe_send_text(instance_wxid, from_wxid, message_text)
                    
                    # 全部完成
                    self.add_log("INFO", f"91卡券处理完成，共发送 {total_cards} 张卡密")
                    
                except requests.RequestException as e:
                    import traceback
                    self.add_log("ERROR", f"请求91卡券API失败: {e}\n{traceback.format_exc()}")
                    self._safe_send_text(instance_wxid, from_wxid, f"网络请求失败: {e}")
                except Exception as e:
                    import traceback
                    self.add_log("ERROR", f"处理91卡券时出错: {e}\n{traceback.format_exc()}")
                    self._safe_send_text(instance_wxid, from_wxid, f"处理失败: {e}")
            
            # 启动子线程
            thread = threading.Thread(target=fetch_and_send)
            thread.daemon = True
            thread.start()
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理91卡券链接时出错: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
    
    def _safe_send_text(self, wxid, to_wxid, content):
        """安全发送文本消息（带异常保护和延迟）"""
        try:
            # 添加延迟防止回复过快
            time.sleep(0.5)
            
            if not self._bridge:
                self.add_log("ERROR", "Bridge未初始化，无法发送消息")
                return False
            
            result = self.send_text(wxid, to_wxid, content)
            if not result.get("success"):
                self.add_log("ERROR", f"发送消息失败: {result.get('error')}")
                return False
            return True
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"发送消息异常: {e}\n{traceback.format_exc()}")
            return False
    
    # ==================== 管理员相关方法 ====================
    
    def _check_admin_permission(self, wxid):
        """
        检查是否是管理员或超级管理员
        
        Returns:
            tuple: (is_admin: bool, admin_type: str)
        """
        if wxid == self.super_admin:
            return True, "超级管理员"
        elif wxid in self.admins:
            return True, "管理员"
        else:
            return False, None
    
    def _handle_set_admin(self, instance_wxid, from_wxid, content):
        """处理设置管理员命令（仅超级管理员可用）"""
        try:
            # 只有超级管理员可以设置管理员
            if from_wxid != self.super_admin:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 权限不足，只有超级管理员可以设置管理员")
                return
            
            # 解析命令：设置管理员@wxid
            parts = content.split('@')
            if len(parts) != 2:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 命令格式错误\n\n正确格式：\n设置管理员@wxid")
                return
            
            target_wxid = parts[1].strip()
            
            if not target_wxid:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 请提供目标用户的wxid")
                return
            
            # 检查是否已经是管理员
            if target_wxid in self.admins:
                self._safe_send_text(instance_wxid, from_wxid, f"ℹ️ 用户 {target_wxid} 已经是管理员")
                return
            
            # 检查是否是超级管理员
            if target_wxid == self.super_admin:
                self._safe_send_text(instance_wxid, from_wxid, f"ℹ️ 用户 {target_wxid} 是超级管理员，无需设置")
                return
            
            # 添加到管理员列表
            self.admins.append(target_wxid)
            
            # 保存到文件
            self._save_admins()
            
            self.add_log("INFO", f"超级管理员 {from_wxid} 设置 {target_wxid} 为管理员")
            
            # 发送成功消息
            msg = f"✅ 管理员设置成功\n\n目标用户: {target_wxid}\n权限: 管理员\n\n当前管理员列表（{len(self.admins)}人）：\n" + "\n".join(self.admins)
            self._safe_send_text(instance_wxid, from_wxid, msg)
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理设置管理员时出错: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, f"设置失败: {str(e)}")
    
    # ==================== 会员管理相关方法 ====================
    
    def _handle_vip_authorization(self, instance_wxid, from_wxid, content):
        """处理会员授权命令"""
        try:
            # 解析命令：月卡授权@wxid 或 季卡授权@wxid 或 年卡授权@wxid
            parts = content.split('@')
            if len(parts) != 2:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 命令格式错误\n\n正确格式：\n月卡授权@wxid\n季卡授权@wxid\n年卡授权@wxid")
                return
            
            vip_type = parts[0]  # "月卡授权" 或 "季卡授权" 或 "年卡授权"
            target_wxid = parts[1].strip()
            
            if not target_wxid:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 请提供目标用户的wxid")
                return
            
            # 计算到期时间
            now = datetime.now()
            if vip_type == "月卡授权":
                # 一个月后
                expire_time = now + timedelta(days=30)
                card_type = "月卡"
            elif vip_type == "季卡授权":
                # 三个月后
                expire_time = now + timedelta(days=90)
                card_type = "季卡"
            elif vip_type == "年卡授权":
                # 一年后
                expire_time = now + timedelta(days=365)
                card_type = "年卡"
            else:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 未知的会员类型")
                return
            
            # 保存会员信息
            self.vip_members[target_wxid] = {
                "type": card_type,
                "expire_time": expire_time.strftime("%Y-%m-%d %H:%M:%S")
            }
            
            # 保存到文件
            self._save_vip_members()
            
            self.add_log("INFO", f"为用户 {target_wxid} 授权{card_type}会员，到期时间: {expire_time.strftime('%Y-%m-%d %H:%M:%S')}")
            
            # 发送成功消息
            msg = f"✅ {card_type}会员授权成功\n\n目标用户: {target_wxid}\n会员类型: {card_type}\n到期时间: {expire_time.strftime('%Y-%m-%d %H:%M:%S')}"
            self._safe_send_text(instance_wxid, from_wxid, msg)
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理会员授权时出错: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, f"授权失败: {str(e)}")
    
    def _check_vip_status(self, wxid):
        """
        检查用户的会员状态
        
        Returns:
            tuple: (is_vip: bool, vip_type: str, expire_time: str)
        """
        if wxid not in self.vip_members:
            return False, None, None
        
        vip_info = self.vip_members[wxid]
        expire_time_str = vip_info.get("expire_time")
        vip_type = vip_info.get("type")
        
        # 检查是否过期
        try:
            expire_time = datetime.strptime(expire_time_str, "%Y-%m-%d %H:%M:%S")
            now = datetime.now()
            
            if now > expire_time:
                # 已过期，移除会员信息
                self.add_log("INFO", f"用户 {wxid} 的{vip_type}会员已过期")
                del self.vip_members[wxid]
                self._save_vip_members()
                return False, None, None
            
            return True, vip_type, expire_time_str
        except Exception as e:
            self.add_log("ERROR", f"检查会员状态时出错: {e}")
            return False, None, None
    
    # ==================== 查返利相关方法 ====================
    
    def _handle_check_rebate(self, instance_wxid, from_wxid):
        """处理查返利命令"""
        try:
            self.add_log("INFO", f"用户 {from_wxid} 请求查返利")
            
            # 保存用户会话信息
            self.user_sessions[from_wxid] = {
                "type": "rebate_wait_token",
                "instance_wxid": instance_wxid
            }
            
            # 发送提示消息
            self._safe_send_text(instance_wxid, from_wxid, "请发送您的美团联盟Token：")
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理查返利命令时出错: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
    
    def _handle_rebate_token_input(self, instance_wxid, from_wxid, content):
        """处理用户输入的token URL"""
        try:
            self.add_log("INFO", f"用户 {from_wxid} 发送token URL: {content[:100]}")
            
            # 从URL中提取token和userid
            # 示例：https://i.meituan.com/mttouch/page/account?cevent=imt%2Fhomepage%2Fmine&userId=2262417476&token=AgHVJRdiHGGvYYiX8LJw1KkgEQdhxIPZho-Fz80L_44sZLjzCOcfnMhYwLwkKsx7M5WfBG6axuPjKAAAAAC3LQAA5EBjkH6O2uCkTrxm-8IsszYc5UiT6dQMNnX1bcHuk37ECJhgeZ3f1Wr6-CsvcUes
            
            # 提取userId
            userid_match = re.search(r'userId=(\d+)', content)
            # 提取token（支持URL编码的&符号）
            token_match = re.search(r'token=([^&\s]+)', content)
            
            if not userid_match or not token_match:
                self.add_log("WARNING", f"无法从URL中提取token或userid: {content[:100]}")
                self._safe_send_text(instance_wxid, from_wxid, "❌ 无法识别Token链接，请确认链接格式是否正确\n\n正确格式示例：\nhttps://i.meituan.com/mttouch/page/account?userId=xxx&token=xxx")
                # 不清除会话，允许重试
                return
            
            userid = userid_match.group(1)
            token = token_match.group(1)
            
            self.add_log("INFO", f"成功提取 userid={userid}, token={token[:50]}...")
            
            # 更新会话状态
            self.user_sessions[from_wxid] = {
                "type": "rebate_wait_order",
                "instance_wxid": instance_wxid,
                "token": token,
                "userid": userid
            }
            
            # 发送提示消息
            self._safe_send_text(instance_wxid, from_wxid, "✅ Token提取成功\n\n请发送目标订单号：")
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理token输入时出错: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
            # 清除会话
            if from_wxid in self.user_sessions:
                del self.user_sessions[from_wxid]
    
    def _handle_rebate_order_input(self, instance_wxid, from_wxid, content):
        """处理用户输入的订单号（支持批量）"""
        try:
            # 检查是否是退出命令
            if content.strip().lower() == "q":
                self.add_log("INFO", f"用户 {from_wxid} 退出查返利")
                # 清除会话
                if from_wxid in self.user_sessions:
                    del self.user_sessions[from_wxid]
                self._safe_send_text(instance_wxid, from_wxid, "✅ 已退出查返利功能")
                return
            
            # 解析订单号（支持多行）
            lines = content.strip().split('\n')
            order_ids = []
            
            for line in lines:
                line = line.strip()
                if line and line.isdigit():
                    order_ids.append(line)
                elif line:  # 非空但不是纯数字
                    self.add_log("WARNING", f"订单号格式错误: {line}")
                    self._safe_send_text(instance_wxid, from_wxid, f"❌ 订单号格式错误: {line}\n请发送纯数字订单号\n\n示例：5008031632996608598\n批量查询请每行一个订单号\n\n退出查询请输入 \"q\"")
                    return
            
            if not order_ids:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 未检测到有效的订单号\n\n示例：5008031632996608598\n批量查询请每行一个订单号\n\n退出查询请输入 \"q\"")
                return
            
            is_batch = len(order_ids) > 1
            self.add_log("INFO", f"用户 {from_wxid} 发送订单号: {'批量' if is_batch else '单个'}, 数量: {len(order_ids)}")
            
            # 获取会话信息
            session = self.user_sessions.get(from_wxid)
            if not session:
                self._safe_send_text(instance_wxid, from_wxid, "会话已过期，请重新发送\"查返利\"命令")
                return
            
            token = session.get("token")
            userid = session.get("userid")
            
            # 检查会员状态
            is_vip, vip_type, expire_time = self._check_vip_status(from_wxid)
            
            # 检查权限并获取用户信息（需要用到余额）
            has_permission, error_msg, user = self._check_user_permission(from_wxid)
            if not has_permission:
                # 清除会话
                if from_wxid in self.user_sessions:
                    del self.user_sessions[from_wxid]
                self._safe_send_text(instance_wxid, from_wxid, error_msg)
                return
            
            rebate_cost = 0.0  # 查返利费用
            if is_vip:
                # 会员用户不受次数限制
                self.add_log("INFO", f"用户 {from_wxid} 是{vip_type}会员，不受次数限制")
                remaining = 99999  # 设置一个很大的数，表示不限制
                daily_usage = 0  # 会员不计次数，设为0仅用于日志
            else:
                # 非会员用户，检查今日查询次数（10次免费，超过10次每次扣0.04元）
                daily_usage = self._get_daily_usage(from_wxid, usage_type="rebate")
                remaining = 99999  # 允许无限次查询，但超过10次需要扣费
                
                if daily_usage >= 10:
                    # 超过10次免费额度，需要扣费
                    rebate_cost = 0.04 * len(order_ids)  # 按订单数量计费
                    if user["balance"] < rebate_cost:
                        self.add_log("WARNING", f"用户 {from_wxid} 余额不足，当前余额: {user['balance']:.2f}元，需要: {rebate_cost:.2f}元")
                        # 清除会话
                        if from_wxid in self.user_sessions:
                            del self.user_sessions[from_wxid]
                        self._safe_send_text(
                            instance_wxid, 
                            from_wxid, 
                            f"❌ 您今日已查询{daily_usage}次（超过10次免费额度）\n本次查询{len(order_ids)}个订单需扣费: {rebate_cost:.2f}元\n当前余额: {user['balance']:.2f}元\n余额不足，请充值后再试！"
                        )
                        return
                    self.add_log("INFO", f"用户 {from_wxid} 今日已查询{daily_usage}次（超过10次免费额度），本次查询{len(order_ids)}个订单需扣费 {rebate_cost:.2f}元")
                else:
                    self.add_log("INFO", f"用户 {from_wxid} 今日已查询{daily_usage}次，在免费额度内")
            
            # 检查是否超过剩余次数
            orders_to_query = order_ids[:remaining]  # 只查询剩余次数允许的订单
            orders_skipped = order_ids[remaining:]   # 跳过的订单
            
            if orders_skipped:
                self.add_log("WARNING", f"用户 {from_wxid} 订单数量({len(order_ids)})超过剩余次数({remaining})，跳过{len(orders_skipped)}个订单")
            
            # 根据会员状态记录不同的日志
            if is_vip:
                self.add_log("INFO", f"开始查询订单返利信息: 数量={len(orders_to_query)}, userid={userid}, {vip_type}会员不限次数")
            else:
                self.add_log("INFO", f"开始查询订单返利信息: 数量={len(orders_to_query)}, userid={userid}, 今日已查询{daily_usage}次")
            
            # 不清除会话，保持状态以支持连续查询
            
            # 发送查询中提示
            if is_batch:
                self._safe_send_text(instance_wxid, from_wxid, f"🔍 正在批量查询返利信息（共{len(orders_to_query)}个订单），请稍候...")
            else:
                self._safe_send_text(instance_wxid, from_wxid, "🔍 正在查询返利信息，请稍候...")
            
            # 在子线程中查询（避免阻塞）
            def query_rebate():
                try:
                    query_results = []  # 存储查询结果
                    has_error = False
                    error_message = ""
                    
                    # 批量查询每个订单
                    for order_id in orders_to_query:
                        # 优先使用本地API接口，出错时使用备用接口
                        try:
                            result = api_get_mt_order_rebate_info(order_id, token, userid)
                            self.add_log("INFO", f"订单 {order_id} 使用本地API查询")
                        except Exception as e:
                            self.add_log("WARNING", f"本地API查询失败: {e}，使用备用接口")
                            try:
                                result = get_mt_order_rebate_info(order_id, token, userid)
                                self.add_log("INFO", f"订单 {order_id} 使用备用接口查询")
                            except Exception as e2:
                                self.add_log("ERROR", f"备用接口也失败: {e2}")
                                result = {"error": f"查询失败: {str(e2)}"}
                        
                        self.add_log("INFO", f"订单 {order_id} 返利查询结果: {str(result)[:200]}")
                        
                        # 判断返回结果类型
                        if isinstance(result, dict):
                            # 检查是否遇到风控 (yodaCode: 406)
                            if result.get('yodaCode') == 406:
                                self.add_log("ERROR", f"订单 {order_id} 查询遇到风控: {result.get('msg', '未知')}")
                                
                                # 提取风控信息
                                general_page_url = result.get('customData', {}).get('generalPageUrl', '')
                                request_code = result.get('customData', {}).get('requestCode', '')
                                risk_level = result.get('customData', {}).get('riskLevel', '')
                                msg = result.get('msg', '未知错误')
                                
                                # 发送风控通知给管理员
                                admin_wxid = "wxid_3intiqznkov222"
                                notification_msg = f"🚨 查返利风控通知\n订单号: {order_id}\n用户: {from_wxid}\n风控消息: {msg}\nriskLevel: {risk_level}\nrequestCode: {request_code}"
                                if general_page_url:
                                    notification_msg += f"\n验证URL:\n{general_page_url}"
                                
                                self._safe_send_text(instance_wxid, admin_wxid, notification_msg)
                                self.add_log("INFO", f"已通知管理员查返利风控情况，订单: {order_id}")
                                
                                # 作为错误处理，停止查询
                                has_error = True
                                error_message = f"订单 {order_id} 遇到风控: {msg}"
                                break
                            
                            # 检查是否有错误
                            if "error" in result:
                                self.add_log("ERROR", f"查询返利接口错误: {result['error']}")
                                has_error = True
                                error_message = result['error']
                                break
                            
                            # 检查data字段
                            data = result.get("data", [])
                            if data and len(data) > 0:
                                # data不是空列表，说明订单走了返利链接
                                self.add_log("INFO", f"订单 {order_id} 已走返利链接")
                                
                                # 提取详细信息（供会员用户查看）
                                detail_info = None
                                if data[0].get("detailList") and len(data[0]["detailList"]) > 0:
                                    detail = data[0]["detailList"][0]
                                    detail_info = {
                                        "orderViewId": detail.get("orderViewId", ""),
                                        "commissionFee": detail.get("commissionFee", ""),
                                        "orderStatus": detail.get("orderStatus", ""),
                                        "orderPayTime": detail.get("orderPayTime", ""),
                                        "cityName": detail.get("cityName", ""),
                                        "consumeCityName": detail.get("consumeCityName", "")
                                    }
                                
                                query_results.append({
                                    "order_id": order_id, 
                                    "status": "success",
                                    "detail": detail_info
                                })
                            else:
                                # data是空列表，说明订单未走返利链接
                                self.add_log("INFO", f"订单 {order_id} 未走返利链接")
                                query_results.append({"order_id": order_id, "status": "failed", "detail": None})
                        
                        elif isinstance(result, str):
                            # 返回的是字符串（可能是403错误）
                            if "403" in result:
                                self.add_log("WARNING", f"Token可能已失效: {result[:100]}")
                                self._safe_send_text(instance_wxid, from_wxid, "❌ Token可能已失效，请重新获取。")
                                # 清除会话
                                if from_wxid in self.user_sessions:
                                    del self.user_sessions[from_wxid]
                                return
                            else:
                                self.add_log("ERROR", f"查询返回异常文本: {result[:200]}")
                                has_error = True
                                error_message = f"返回异常: {result[:100]}"
                                break
                        else:
                            self.add_log("ERROR", f"查询返回未知类型: {type(result)}")
                            has_error = True
                            error_message = "返回数据格式异常"
                            break
                        
                        # 批量查询时，添加短暂延迟避免请求过快
                        if is_batch and order_id != orders_to_query[-1]:
                            time.sleep(0.5)
                    
                    # 如果有错误，提示用户
                    if has_error:
                        self._safe_send_text(instance_wxid, from_wxid, f"查询失败: {error_message}\n\n继续查询请发送订单号，退出请输入 \"q\"\n=======")
                        return
                    
                    # 查询成功，处理计费和次数统计
                    is_vip_current, vip_type_current, expire_time_current = self._check_vip_status(from_wxid)
                    
                    if not is_vip_current:
                        # 非会员用户，增加使用次数（按订单数量）
                        for _ in range(len(orders_to_query)):
                            self._increment_daily_usage(from_wxid, usage_type="rebate")
                        
                        # 如果需要扣费，执行扣费
                        if rebate_cost > 0:
                            if self._deduct_balance(from_wxid, rebate_cost):
                                self.add_log("INFO", f"已扣除用户 {from_wxid} 余额 {rebate_cost:.2f}元（查询{len(orders_to_query)}个订单）")
                            else:
                                self.add_log("ERROR", f"扣除用户 {from_wxid} 余额失败")
                                self._safe_send_text(instance_wxid, from_wxid, "⚠️ 扣费失败，请联系管理员")
                    
                    # 检查会员状态（用于显示）
                    is_vip, vip_type, expire_time = self._check_vip_status(from_wxid)
                    
                    # 构建回复消息
                    if is_vip:
                        # 会员用户，显示会员信息
                        usage_text = f"尊敬的{vip_type}会员\n您不受此功能次数限制\n您的授权到期时间为：{expire_time}"
                        self.add_log("INFO", f"用户 {from_wxid} 是{vip_type}会员")
                    else:
                        # 普通用户，显示次数信息
                        current_usage = self._get_daily_usage(from_wxid, usage_type="rebate")
                        if current_usage <= 10:
                            remaining = 10 - current_usage
                            usage_text = f"今日已查询 {current_usage}/10 次，剩余 {remaining} 次免费额度\n超过10次后每次扣0.04元/订单"
                        else:
                            usage_text = f"今日已查询 {current_usage} 次（超过10次免费额度）\n本次查询已扣费: {rebate_cost:.2f}元"
                        self.add_log("INFO", f"用户 {from_wxid} 今日查询次数: {current_usage}次")
                    
                    if is_batch:
                        # 批量查询结果
                        # 会员用户生成Excel文件
                        if is_vip and vip_type in ["月卡", "季卡", "年卡"]:
                            # 生成Excel文件
                            excel_file_path = self._save_rebate_results_to_excel(query_results)
                            
                            if excel_file_path:
                                # 发送提示消息
                                self._safe_send_text(instance_wxid, from_wxid, "📊 正在生成Excel文件，请稍候...")
                                
                                # 发送Excel文件
                                result = self.send_file(instance_wxid, from_wxid, excel_file_path)
                                
                                if result.get("success"):
                                    # 发送成功，构建简要文字总结
                                    success_count = sum(1 for item in query_results if item["status"] == "success")
                                    failed_count = len(query_results) - success_count
                                    
                                    summary_msg = f"✅ 批量查询完成！\n\n已走返利: {success_count} 个订单\n未走返利: {failed_count} 个订单\n\nExcel文件已发送，请查看详细信息。"
                                    
                                    # 添加跳过的订单提示
                                    if orders_skipped:
                                        summary_msg += f"\n\n⚠️ 由于次数不足，跳过了{len(orders_skipped)}个订单"
                                    
                                    # 添加会员信息和继续提示
                                    summary_msg += f"\n\n{usage_text}\n\n继续查询请发送订单号，退出请输入 \"q\"\n======="
                                    
                                    self._safe_send_text(instance_wxid, from_wxid, summary_msg)
                                    
                                    # 删除临时文件（如果配置允许）
                                    if self.delete_after_send:
                                        try:
                                            os.remove(excel_file_path)
                                            self.add_log("INFO", f"已删除临时Excel文件: {excel_file_path}")
                                        except Exception as e:
                                            self.add_log("WARNING", f"删除Excel文件失败: {e}")
                                else:
                                    error_msg = result.get("error", "未知错误")
                                    self._safe_send_text(instance_wxid, from_wxid, f"Excel文件发送失败: {error_msg}")
                            else:
                                # Excel生成失败，发送文字版结果
                                self._safe_send_text(instance_wxid, from_wxid, "Excel文件生成失败，为您展示文字版结果：")
                                
                                result_lines = ["批量查询结果："]
                                for item in query_results:
                                    order_id = item["order_id"]
                                    status = item["status"]
                                    status_text = "走返利成功" if status == "success" else "走返利失败"
                                    result_lines.append(f"{order_id}-{status_text}")
                                
                                reply_msg = "\n".join(result_lines)
                                
                                if orders_skipped:
                                    reply_msg += f"\n\n⚠️ 由于次数不足，跳过了{len(orders_skipped)}个订单"
                                
                                reply_msg += f"\n\n{usage_text}\n\n继续查询请发送订单号，退出请输入 \"q\"\n======="
                                self._safe_send_text(instance_wxid, from_wxid, reply_msg)
                        else:
                            # 普通用户，显示文字版结果
                            result_lines = ["批量查询结果如下："]
                            for item in query_results:
                                order_id = item["order_id"]
                                status = item["status"]
                                status_text = "走返利成功" if status == "success" else "走返利失败"
                                result_lines.append(f"{order_id}-{status_text}")
                            
                            reply_msg = "\n".join(result_lines)
                            
                            # 添加跳过的订单提示
                            if orders_skipped:
                                reply_msg += f"\n\n⚠️ 由于次数不足，跳过了{len(orders_skipped)}个订单"
                            
                            # 添加次数信息提示
                            if remaining > 0:
                                reply_msg += f"\n\n{usage_text}\n\n继续查询请发送订单号，退出请输入 \"q\"\n======="
                            else:
                                reply_msg += f"\n\n今日查询次数已用完（{current_usage}/10）"
                                # 清除会话
                                if from_wxid in self.user_sessions:
                                    del self.user_sessions[from_wxid]
                            
                            self._safe_send_text(instance_wxid, from_wxid, reply_msg)
                    else:
                        # 单个查询结果
                        status = query_results[0]["status"]
                        
                        # 对会员显示详细信息
                        if is_vip and vip_type in ["月卡", "季卡", "年卡"] and status == "success" and query_results[0].get("detail"):
                            detail = query_results[0]["detail"]
                            # 转换卡券状态
                            order_status = detail.get("orderStatus", "")
                            if order_status == 3:
                                status_desc = "已退款"
                            elif order_status == 4:
                                status_desc = "未核销"
                            elif order_status == 5:
                                status_desc = "已核销"
                            else:
                                status_desc = str(order_status) if order_status else "未知"
                            
                            # 转换下单时间
                            order_pay_time = detail.get("orderPayTime", "")
                            if order_pay_time:
                                try:
                                    if isinstance(order_pay_time, int):
                                        time_str = datetime.fromtimestamp(order_pay_time).strftime("%Y-%m-%d %H:%M:%S")
                                    else:
                                        time_str = str(order_pay_time)
                                except:
                                    time_str = str(order_pay_time)
                            else:
                                time_str = ""
                            
                            result_msg = f"""✅ 查询成功：此订单已成功走上您的返利链接

订单号：{query_results[0]["order_id"]}
推广单号：{detail.get("orderViewId", "")}
佣金：{detail.get("commissionFee", "")}元
卡券状态：{status_desc}
下单时间：{time_str}
下单城市：{detail.get("cityName", "")}
核销城市：{detail.get("consumeCityName", "") if detail.get("consumeCityName") else ""}"""
                        else:
                            if status == "success":
                                result_msg = "✅ 查询成功：此订单已成功走上您的返利链接"
                            else:
                                result_msg = "⚠️ 查询警告：此订单未走上您的返利链接"
                        
                        # 构建回复消息
                        if is_vip:
                            reply_msg = f"{result_msg}\n\n{usage_text}\n\n继续查询请发送订单号，退出请输入 \"q\"\n======="
                        else:
                            if remaining > 0:
                                reply_msg = f"{result_msg}\n\n{usage_text}\n\n继续查询请发送订单号，退出请输入 \"q\"\n======="
                            else:
                                reply_msg = f"{result_msg}\n\n今日查询次数已用完（{current_usage}/10）"
                                # 清除会话
                                if from_wxid in self.user_sessions:
                                    del self.user_sessions[from_wxid]
                        
                        # 单个查询结果发送消息
                        self._safe_send_text(instance_wxid, from_wxid, reply_msg)
                        
                except Exception as e:
                    import traceback
                    self.add_log("ERROR", f"查询订单返利时出错: {e}\n{traceback.format_exc()}")
                    self._safe_send_text(instance_wxid, from_wxid, f"查询失败: {str(e)}\n\n继续查询请发送订单号，退出请输入 \"q\"\n=======")
            
            # 启动子线程
            thread = threading.Thread(target=query_rebate)
            thread.daemon = True
            thread.start()
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理订单号输入时出错: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
            # 清除会话
            if from_wxid in self.user_sessions:
                del self.user_sessions[from_wxid]
    
    def _check_user_permission(self, from_wxid):
        """
        检查用户权限
        
        Returns:
            tuple: (bool, str, dict) - (是否有权限, 错误信息, 用户信息)
        """
        # 查找用户
        user = None
        for u in self.users:
            if u["wxid"] == from_wxid:
                user = u
                break
        
        if not user:
            return False, "您还未开户，请先发送\"聚合开户\"进行开户\n\n如需帮助请发送\"教程\"查看使用教程", None
        
        if user["status"] != "正常":
            return False, f"您的账户状态异常（{user['status']}），无法使用此功能", user
        
        return True, "", user
    
    def _load_city_dict(self):
        """加载城市字典"""
        try:
            city_dict_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "city_dict.json")
            if os.path.exists(city_dict_path):
                with open(city_dict_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.city_dict = data.get("city_dict", {})
                self.add_log("INFO", f"城市字典加载成功，共{len(self.city_dict)}个城市")
            else:
                self.add_log("WARNING", f"城市字典文件不存在: {city_dict_path}")
        except Exception as e:
            self.add_log("ERROR", f"加载城市字典时出错: {e}")
    
    def _get_city_id(self, city_name):
        """根据城市名获取城市ID"""
        city_name = city_name.strip()
        if city_name in self.city_dict:
            return self.city_dict[city_name]["id"], self.city_dict[city_name]["full_name"]
        return None, None
    
    def _get_daily_usage(self, wxid, usage_type="crawl"):
        """
        获取用户今日使用次数
        
        Args:
            wxid: 用户ID
            usage_type: 使用类型，"crawl"=爬门店, "rebate"=查返利
        
        Returns:
            int: 今日使用次数
        """
        today = datetime.now().strftime("%Y-%m-%d")
        
        for user in self.users:
            if user["wxid"] == wxid:
                # 根据类型选择不同的字段
                usage_key = "daily_usage" if usage_type == "crawl" else "daily_rebate_usage"
                
                # 初始化每日使用记录
                if usage_key not in user:
                    user[usage_key] = {"date": today, "count": 0}
                
                # 检查日期是否为今天，如果不是则重置
                if user[usage_key]["date"] != today:
                    user[usage_key] = {"date": today, "count": 0}
                    self.set_config("users", self.users)
                
                return user[usage_key]["count"]
        
        return 0
    
    def _increment_daily_usage(self, wxid, usage_type="crawl"):
        """
        增加用户今日使用次数
        
        Args:
            wxid: 用户ID
            usage_type: 使用类型，"crawl"=爬门店, "rebate"=查返利
        
        Returns:
            bool: 是否成功
        """
        today = datetime.now().strftime("%Y-%m-%d")
        
        for user in self.users:
            if user["wxid"] == wxid:
                # 根据类型选择不同的字段
                usage_key = "daily_usage" if usage_type == "crawl" else "daily_rebate_usage"
                
                # 初始化每日使用记录
                if usage_key not in user:
                    user[usage_key] = {"date": today, "count": 0}
                
                # 检查日期是否为今天，如果不是则重置
                if user[usage_key]["date"] != today:
                    user[usage_key] = {"date": today, "count": 0}
                
                # 增加计数
                user[usage_key]["count"] += 1
                self.set_config("users", self.users)
                return True
        
        return False
    
    def _deduct_balance(self, wxid, amount):
        """
        扣除用户余额
        
        Returns:
            bool: 是否扣除成功
        """
        for user in self.users:
            if user["wxid"] == wxid:
                if user["balance"] < amount:
                    return False
                user["balance"] -= amount
                self.set_config("users", self.users)
                # 使用QTimer在主线程中更新UI
                self._safe_update_user_table()
                return True
        return False
    
    def _add_balance(self, wxid, amount):
        """添加用户余额"""
        for user in self.users:
            if user["wxid"] == wxid:
                user["balance"] += amount
                self.set_config("users", self.users)
                # 使用QTimer在主线程中更新UI
                self._safe_update_user_table()
                return True
        return False
    
    def _ban_user(self, wxid):
        """封禁用户"""
        for user in self.users:
            if user["wxid"] == wxid:
                user["status"] = "禁用"
                self.set_config("users", self.users)
                # 使用QTimer在主线程中更新UI
                self._safe_update_user_table()
                return True
        return False
    
    def _unban_user(self, wxid):
        """解封用户"""
        for user in self.users:
            if user["wxid"] == wxid:
                user["status"] = "正常"
                self.set_config("users", self.users)
                # 使用QTimer在主线程中更新UI
                self._safe_update_user_table()
                return True
        return False
    
    # ==================== 查企业返利相关方法 ====================
    
    def _handle_check_enterprise_rebate(self, instance_wxid, from_wxid):
        """处理查企业返利命令"""
        try:
            self.add_log("INFO", f"用户 {from_wxid} 请求查企业返利")
            
            # 保存用户会话信息
            self.user_sessions[from_wxid] = {
                "type": "enterprise_rebate_wait_sk",
                "instance_wxid": instance_wxid
            }
            
            # 发送提示消息
            self._safe_send_text(instance_wxid, from_wxid, "请发送您公司的SK：\n（输入\"q\"退出会话）")
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理查企业返利命令时出错: {e}\n{traceback.format_exc()}")
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
    
    def _handle_enterprise_sk_input(self, instance_wxid, from_wxid, sk):
        """处理用户输入的SK"""
        try:
            # 检查是否是退出命令
            if sk.strip().lower() == "q":
                self.add_log("INFO", f"用户 {from_wxid} 退出查企业返利")
                if from_wxid in self.user_sessions:
                    del self.user_sessions[from_wxid]
                self._safe_send_text(instance_wxid, from_wxid, "✅ 已退出查企业返利功能")
                return
            
            sk = sk.strip()
            self.add_log("INFO", f"用户 {from_wxid} 发送SK: {sk[:20]}...")
            
            # 查找SK对应的企业配置
            if sk not in self.enterprise_configs:
                self.add_log("WARNING", f"未找到SK对应的企业配置: {sk}")
                self._safe_send_text(instance_wxid, from_wxid, "❌ 未找到对应的企业配置，请检查SK是否正确\n\n继续输入请发送SK，退出请输入 \"q\"")
                return
            
            enterprise_config = self.enterprise_configs[sk]
            enterprise_name = enterprise_config.get("name", "未知企业")
            
            # 更新会话信息
            self.user_sessions[from_wxid] = {
                "type": "enterprise_rebate_wait_order",
                "instance_wxid": instance_wxid,
                "sk": sk,
                "app_key": enterprise_config["app_key"],
                "app_secret": enterprise_config["app_secret"],
                "enterprise_name": enterprise_name
            }
            
            self.add_log("INFO", f"用户 {from_wxid} SK验证成功，企业: {enterprise_name}")
            
            # 发送提示消息
            self._safe_send_text(
                instance_wxid, 
                from_wxid, 
                f"尊敬的{enterprise_name}，请发送需要查返利的订单号：\n（多个订单号请换行）\n（输入\"q\"退出会话）"
            )
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理SK输入时出错: {e}\n{traceback.format_exc()}")
            if from_wxid in self.user_sessions:
                del self.user_sessions[from_wxid]
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
    
    def _handle_enterprise_order_input(self, instance_wxid, from_wxid, content):
        """处理企业返利订单号输入（支持批量）"""
        try:
            # 检查是否是退出命令
            if content.strip().lower() == "q":
                self.add_log("INFO", f"用户 {from_wxid} 退出查企业返利")
                if from_wxid in self.user_sessions:
                    del self.user_sessions[from_wxid]
                self._safe_send_text(instance_wxid, from_wxid, "✅ 已退出查企业返利功能")
                return
            
            # 解析订单号（支持多行）
            lines = content.strip().split('\n')
            order_ids = []
            
            for line in lines:
                line = line.strip()
                if line and line.isdigit():
                    order_ids.append(line)
                elif line:  # 非空但不是纯数字
                    self.add_log("WARNING", f"订单号格式错误: {line}")
                    self._safe_send_text(instance_wxid, from_wxid, f"❌ 订单号格式错误: {line}\n请发送纯数字订单号\n\n示例：5008031697162667606\n批量查询请每行一个订单号\n\n退出查询请输入 \"q\"")
                    return
            
            if not order_ids:
                self._safe_send_text(instance_wxid, from_wxid, "❌ 未检测到有效的订单号\n\n示例：5008031697162667606\n批量查询请每行一个订单号\n\n退出查询请输入 \"q\"")
                return
            
            is_batch = len(order_ids) > 1
            self.add_log("INFO", f"用户 {from_wxid} 发送订单号: {'批量' if is_batch else '单个'}, 数量: {len(order_ids)}")
            
            # 获取会话信息
            session = self.user_sessions.get(from_wxid)
            if not session:
                self._safe_send_text(instance_wxid, from_wxid, "会话已过期，请重新发送\"查企业返利\"命令")
                return
            
            app_key = session.get("app_key")
            app_secret = session.get("app_secret")
            enterprise_name = session.get("enterprise_name")
            
            self.add_log("INFO", f"开始查询企业返利: 企业={enterprise_name}, 数量={len(order_ids)}")
            
            # 发送查询中提示
            if is_batch:
                self._safe_send_text(instance_wxid, from_wxid, f"🔍 正在批量查询返利信息（共{len(order_ids)}个订单），请稍候...")
            else:
                self._safe_send_text(instance_wxid, from_wxid, "🔍 正在查询返利信息，请稍候...")
            
            # 在子线程中查询（避免阻塞）
            def query_enterprise_orders():
                try:
                    # 整理查询结果
                    query_results = []
                    
                    # 循环查询每个订单号（API要求orderId必须是字符串，不支持数组）
                    for idx, order_id in enumerate(order_ids, start=1):
                        try:
                            self.add_log("INFO", f"正在查询订单 {idx}/{len(order_ids)}: {order_id}")
                            
                            # 调用query_orders接口
                            params = {
                                "limit": 20,
                                "queryTimeType": 1,
                                "page": 1,
                                "platform": 2,
                                "businessLine": [1, 2],
                                "orderId": order_id  # 必须是字符串类型
                            }
                            
                            result = query_orders(
                                app_key=app_key,
                                app_secret=app_secret,
                                query_params=params
                            )
                            
                            self.add_log("INFO", f"订单 {order_id} 查询返回: {str(result)[:200]}")
                            
                            # 检查返回结果
                            if result.get("code") != 0:
                                error_msg = result.get("message", "未知错误")
                                self.add_log("ERROR", f"查询订单 {order_id} 失败: {error_msg}")
                                # 记录失败，但继续查询其他订单
                                query_results.append({
                                    "plain_order_id": order_id,  # 明文订单号（用户发送的）
                                    "order_id": order_id,
                                    "pay_price": "0",
                                    "profit": "0",
                                    "cpa_profit": "0",
                                    "product_name": f"查询失败: {error_msg}",
                                    "status": "查询失败",
                                    "pay_time": "",
                                    "city_name": "",
                                    "commission_rate": "0",
                                    "detail": None
                                })
                                continue
                            
                            data = result.get("data", {})
                            data_list = data.get("dataList", [])
                            
                            if not data_list:
                                self.add_log("INFO", f"订单 {order_id} 查询结果为空")
                                # 记录为查询结果为空
                                query_results.append({
                                    "plain_order_id": order_id,  # 明文订单号（用户发送的）
                                    "order_id": order_id,
                                    "pay_price": "0",
                                    "profit": "0",
                                    "cpa_profit": "0",
                                    "product_name": "未找到订单信息",
                                    "status": "无数据",
                                    "pay_time": "",
                                    "city_name": "",
                                    "commission_rate": "0",
                                    "detail": None
                                })
                                continue
                            
                            self.add_log("INFO", f"订单 {order_id} 查询到 {len(data_list)} 条记录")
                            
                            # 处理查询结果（可能返回多条记录）
                            for order_data in data_list:
                                try:
                                    # 提取订单信息
                                    returned_order_id = order_data.get("orderId", "")
                                    pay_price = order_data.get("payPrice", "0")
                                    profit = order_data.get("profit", "0")
                                    cpa_profit = order_data.get("cpaProfit", "0")
                                    product_name = order_data.get("productName", "未知商品")
                                    status_code = order_data.get("status", "")
                                    pay_time = order_data.get("payTime", 0)
                                    city_name = order_data.get("cityName", "")
                                    commission_rate = order_data.get("commissionRate", "0")
                                    
                                    # 转换订单状态为中文
                                    status_map = {
                                        "2": "付款",
                                        "3": "完成",
                                        "4": "取消",
                                        "5": "风控",
                                        "6": "结算",
                                        2: "付款",
                                        3: "完成",
                                        4: "取消",
                                        5: "风控",
                                        6: "结算"
                                    }
                                    status = status_map.get(status_code, str(status_code) if status_code else "未知")
                                    
                                    # 时间戳转换
                                    pay_time_str = datetime.fromtimestamp(pay_time).strftime("%Y-%m-%d %H:%M:%S") if pay_time else ""
                                    
                                    # 订单详情
                                    order_details = order_data.get("orderDetail", [])
                                    detail_info = None
                                    if order_details and len(order_details) > 0:
                                        detail = order_details[0]
                                        coupon_status_code = detail.get("couponStatus", "")
                                        
                                        # 转换券状态为中文
                                        coupon_status_map = {
                                            "1": "付款",
                                            "2": "已核销",
                                            "3": "结算",
                                            "4": "失效（含取消或风控）",
                                            1: "付款",
                                            2: "已核销",
                                            3: "结算",
                                            4: "失效（含取消或风控）"
                                        }
                                        coupon_status = coupon_status_map.get(coupon_status_code, str(coupon_status_code) if coupon_status_code else "未知")
                                        
                                        detail_info = {
                                            "couponStatus": coupon_status,
                                            "itemOrderId": detail.get("itemOrderId", ""),
                                            "couponFee": detail.get("couponFee", "0"),
                                            "basicAmount": detail.get("basicAmount", "0")
                                        }
                                    
                                    query_results.append({
                                        "plain_order_id": order_id,  # 明文订单号（用户发送的）
                                        "order_id": returned_order_id,
                                        "pay_price": pay_price,
                                        "profit": profit,
                                        "cpa_profit": cpa_profit,
                                        "product_name": product_name,
                                        "status": status,
                                        "pay_time": pay_time_str,
                                        "city_name": city_name,
                                        "commission_rate": commission_rate,
                                        "detail": detail_info
                                    })
                                    
                                except Exception as e:
                                    self.add_log("WARNING", f"解析订单数据时出错: {e}")
                                    continue
                            
                            # 批量查询时添加短暂延迟避免请求过快
                            if is_batch and idx < len(order_ids):
                                time.sleep(0.3)
                                
                        except Exception as e:
                            self.add_log("ERROR", f"查询订单 {order_id} 时出错: {e}")
                            continue
                    
                    # 构建回复消息
                    if is_batch:
                        # 批量查询
                        if len(query_results) > 5:
                            # 订单数大于5，生成Excel文件并发送
                            self.add_log("INFO", f"批量查询订单数({len(query_results)})大于5，生成Excel文件")
                            
                            # 生成Excel文件
                            excel_path = self._save_enterprise_rebate_results_to_excel(query_results, enterprise_name)
                            
                            if excel_path:
                                self.add_log("INFO", f"Excel文件生成成功: {excel_path}")
                                
                                # 发送简要文字信息
                                summary_msg = f"✅ 【{enterprise_name}】批量查询完成\n\n查询结果：共{len(query_results)}个订单\nExcel文件正在发送，请稍候..."
                                self._safe_send_text(instance_wxid, from_wxid, summary_msg)
                                
                                # 发送Excel文件
                                result = self.send_file(instance_wxid, from_wxid, excel_path)
                                
                                if result.get("success"):
                                    self.add_log("INFO", f"Excel文件发送成功")
                                    self._safe_send_text(instance_wxid, from_wxid, "✅ Excel文件已发送，请查收\n\n继续查询请发送订单号，退出请输入 \"q\"\n=======")
                                    
                                    # 如果设置了发送后删除，则删除文件
                                    if self.delete_after_send:
                                        try:
                                            os.remove(excel_path)
                                            self.add_log("INFO", f"已删除Excel文件: {excel_path}")
                                        except Exception as e:
                                            self.add_log("WARNING", f"删除Excel文件失败: {e}")
                                else:
                                    error_msg = result.get("error", "未知错误")
                                    self.add_log("ERROR", f"Excel文件发送失败: {error_msg}")
                                    self._safe_send_text(instance_wxid, from_wxid, f"❌ Excel文件发送失败: {error_msg}\n\n继续查询请发送订单号，退出请输入 \"q\"\n=======")
                            else:
                                # Excel生成失败，发送文字版结果
                                self._safe_send_text(instance_wxid, from_wxid, "Excel文件生成失败，为您展示文字版结果：")
                                
                                result_lines = [f"📊 【{enterprise_name}】批量查询结果（共{len(query_results)}个订单）：\n"]
                                for idx, item in enumerate(query_results, start=1):
                                    result_lines.append(f"\n🔹 订单 {idx}")
                                    result_lines.append(f"明文订单号：{item['plain_order_id']}")
                                    result_lines.append(f"订单号：{item['order_id']}")
                                    result_lines.append(f"商品：{item['product_name']}")
                                    result_lines.append(f"佣金：{item['profit']}元")
                                
                                result_lines.append("\n===================")
                                result_lines.append("\n继续查询请发送订单号，退出请输入 \"q\"\n=======")
                                
                                reply_msg = "\n".join(result_lines)
                                self._safe_send_text(instance_wxid, from_wxid, reply_msg)
                        else:
                            # 订单数<=5，发送文字版汇总信息
                            result_lines = [f"📊 【{enterprise_name}】批量查询结果（共{len(query_results)}个订单）：\n"]
                            
                            for idx, item in enumerate(query_results, start=1):
                                result_lines.append(f"\n🔹 订单 {idx}")
                                result_lines.append(f"明文订单号：{item['plain_order_id']}")
                                result_lines.append(f"订单号：{item['order_id']}")
                                result_lines.append(f"商品名称：{item['product_name']}")
                                result_lines.append(f"支付金额：{item['pay_price']}元")
                                result_lines.append(f"佣金：{item['profit']}元")
                                result_lines.append(f"CPA佣金：{item['cpa_profit']}元")
                                result_lines.append(f"佣金比例：{item['commission_rate']}%")
                                result_lines.append(f"城市：{item['city_name']}")
                                result_lines.append(f"支付时间：{item['pay_time']}")
                                result_lines.append(f"订单状态：{item['status']}")
                            
                            result_lines.append("\n===================")
                            result_lines.append("\n继续查询请发送订单号，退出请输入 \"q\"\n=======")
                            
                            reply_msg = "\n".join(result_lines)
                            self._safe_send_text(instance_wxid, from_wxid, reply_msg)
                    else:
                        # 单个查询，发送详细信息
                        item = query_results[0]
                        msg_lines = [f"✅ 【{enterprise_name}】订单返利查询成功\n"]
                        msg_lines.append(f"明文订单号：{item['plain_order_id']}")
                        msg_lines.append(f"订单号：{item['order_id']}")
                        msg_lines.append(f"商品名称：{item['product_name']}")
                        msg_lines.append(f"支付金额：{item['pay_price']}元")
                        msg_lines.append(f"佣金：{item['profit']}元")
                        msg_lines.append(f"CPA佣金：{item['cpa_profit']}元")
                        msg_lines.append(f"佣金比例：{item['commission_rate']}%")
                        msg_lines.append(f"城市：{item['city_name']}")
                        msg_lines.append(f"支付时间：{item['pay_time']}")
                        msg_lines.append(f"订单状态：{item['status']}")
                        
                        if item['detail']:
                            msg_lines.append(f"\n详细信息：")
                            msg_lines.append(f"券状态：{item['detail']['couponStatus']}")
                            msg_lines.append(f"子订单号：{item['detail']['itemOrderId']}")
                            msg_lines.append(f"券面额：{item['detail']['couponFee']}元")
                            msg_lines.append(f"基础金额：{item['detail']['basicAmount']}元")
                        
                        msg_lines.append("\n继续查询请发送订单号，退出请输入 \"q\"\n=======")
                        
                        reply_msg = "\n".join(msg_lines)
                        self._safe_send_text(instance_wxid, from_wxid, reply_msg)
                    
                except Exception as e:
                    import traceback
                    self.add_log("ERROR", f"查询企业返利时出错: {e}\n{traceback.format_exc()}")
                    self._safe_send_text(instance_wxid, from_wxid, f"查询失败: {str(e)}\n\n继续查询请发送订单号，退出请输入 \"q\"\n=======")
            
            # 启动子线程
            thread = threading.Thread(target=query_enterprise_orders)
            thread.daemon = True
            thread.start()
            
        except Exception as e:
            import traceback
            self.add_log("ERROR", f"处理企业返利订单号时出错: {e}\n{traceback.format_exc()}")
            if from_wxid in self.user_sessions:
                del self.user_sessions[from_wxid]
            self._safe_send_text(instance_wxid, from_wxid, "处理失败，请稍后重试")
    
    def _save_enterprise_rebate_results_to_excel(self, query_results, enterprise_name):
        """
        将企业返利查询结果保存为Excel文件
        
        Args:
            query_results: 查询结果列表
            enterprise_name: 企业名称
            
        Returns:
            str: 文件路径，失败返回None
        """
        try:
            # 检查openpyxl是否可用
            if not OPENPYXL_AVAILABLE:
                self.add_log("ERROR", "缺少openpyxl库，无法生成Excel文件")
                return None
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "企业返利查询结果"
            
            # 写入表头
            ws.append(["明文订单号", "订单号", "商品名称", "支付金额(元)", "佣金(元)", "CPA佣金(元)", "佣金比例(%)", "城市", "支付时间", "订单状态", "券状态", "子订单号", "券面额(元)", "基础金额(元)"])
            
            # 写入数据
            for item in query_results:
                plain_order_id = item.get("plain_order_id", "")
                order_id = item.get("order_id", "")
                product_name = item.get("product_name", "")
                pay_price = item.get("pay_price", "")
                profit = item.get("profit", "")
                cpa_profit = item.get("cpa_profit", "")
                commission_rate = item.get("commission_rate", "")
                city_name = item.get("city_name", "")
                pay_time = item.get("pay_time", "")
                status = item.get("status", "")
                
                detail = item.get("detail")
                if detail:
                    coupon_status = detail.get("couponStatus", "")
                    item_order_id = detail.get("itemOrderId", "")
                    coupon_fee = detail.get("couponFee", "")
                    basic_amount = detail.get("basicAmount", "")
                    
                    ws.append([
                        plain_order_id,
                        order_id,
                        product_name,
                        pay_price,
                        profit,
                        cpa_profit,
                        commission_rate,
                        city_name,
                        pay_time,
                        status,
                        coupon_status,
                        item_order_id,
                        coupon_fee,
                        basic_amount
                    ])
                else:
                    ws.append([
                        plain_order_id,
                        order_id,
                        product_name,
                        pay_price,
                        profit,
                        cpa_profit,
                        commission_rate,
                        city_name,
                        pay_time,
                        status,
                        "", "", "", ""
                    ])
            
            # 调整列宽
            ws.column_dimensions['A'].width = 25  # 明文订单号
            ws.column_dimensions['B'].width = 25  # 订单号
            ws.column_dimensions['C'].width = 30  # 商品名称
            ws.column_dimensions['D'].width = 12  # 支付金额
            ws.column_dimensions['E'].width = 10  # 佣金
            ws.column_dimensions['F'].width = 12  # CPA佣金
            ws.column_dimensions['G'].width = 12  # 佣金比例
            ws.column_dimensions['H'].width = 15  # 城市
            ws.column_dimensions['I'].width = 20  # 支付时间
            ws.column_dimensions['J'].width = 10  # 订单状态
            ws.column_dimensions['K'].width = 10  # 券状态
            ws.column_dimensions['L'].width = 20  # 子订单号
            ws.column_dimensions['M'].width = 12  # 券面额
            ws.column_dimensions['N'].width = 12  # 基础金额
            
            # 生成时间戳（Unix时间戳）
            timestamp = str(int(time.time()))
            
            # 生成文件名
            filename = f"企业返利查询结果-{enterprise_name}-{timestamp}.xlsx"
            
            # 保存到专门的Excel输出目录
            file_path = os.path.join(self.excel_output_dir, filename)
            
            # 确保目录存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # 保存文件
            wb.save(file_path)
            
            self.add_log("INFO", f"企业返利查询Excel文件已保存到本地: {file_path}")
            
            return file_path
        
        except Exception as e:
            self.add_log("ERROR", f"保存企业返利查询Excel文件时出错: {e}")
            import traceback
            self.add_log("ERROR", f"详细错误: {traceback.format_exc()}")
            return None
    
    # ==================== 美团门店抓取相关方法 ====================
    
    def _handle_miniprogram_message(self, instance_wxid, from_wxid, content):
        """处理小程序链接消息 - 显示功能选择菜单"""
        try:
            # 检查用户权限
            has_permission, error_msg, user = self._check_user_permission(from_wxid)
            if not has_permission:
                self.add_log("WARNING", f"用户无权限: {from_wxid} - {error_msg}")
                self._safe_send_text(instance_wxid, from_wxid, error_msg)
                return
            
            self.add_log("INFO", f"收到小程序链接消息，来自: {from_wxid}")
            
            # 提取美团ID
            mt_id = get_coupon_mt_id(content)
            
            if mt_id == 0 or not mt_id:
                self.add_log("WARNING", "未能从消息中提取到美团ID")
                return
            
            self.add_log("INFO", f"提取到美团ID: {mt_id}")
            
            # 保存用户会话信息
            self.user_sessions[from_wxid] = {
                "type": "mt_menu",
                "mt_id": str(mt_id),
                "content": content,
                "instance_wxid": instance_wxid
            }
            
            # 获取今日使用次数
            daily_usage = self._get_daily_usage(from_wxid)
            
            # 发送功能选择菜单
            menu_msg = (
                "此功能免费，请选择功能：\n"
                "1、获取美团ID\n"
                "2、直接抓取门店\n\n"
                "💡 直接发送功能编号，如发送 \"1\" 或 \"2\"\n\n"
                f"功能2 \"直接抓取门店\" 单日单号限制1次\n"
                f"超过1次将每次扣除0.5￥（触发缓存不扣费）\n"
                f"您今日已使用次数： {daily_usage} "
            )
            self._safe_send_text(instance_wxid, from_wxid, menu_msg)
            
        except Exception as e:
            self.add_log("ERROR", f"处理小程序消息时出错: {e}")
            try:
                self._safe_send_text(instance_wxid, from_wxid, f"处理出错: {str(e)}")
            except:
                pass
    
    def _handle_user_choice(self, instance_wxid, from_wxid, choice):
        """处理用户选择"""
        try:
            # 获取会话信息
            session = self.user_sessions.get(from_wxid)
            if not session:
                return
            
            session_type = session.get("type")
            
            # 处理美团菜单类型的会话
            if session_type == "mt_menu":
                mt_id = session.get("mt_id")
                content = session.get("content")
                
                # 处理用户选择
                if choice == "1":
                    # 选择1：获取美团ID
                    self.add_log("INFO", f"用户 {from_wxid} 选择：获取美团ID")
                    self._safe_send_text(instance_wxid, from_wxid, f"✅ 美团ID：{mt_id}")
                    # 清除会话
                    del self.user_sessions[from_wxid]
                    
                elif choice == "2":
                    # 选择2：直接抓取门店
                    self.add_log("INFO", f"用户 {from_wxid} 选择：直接抓取门店")
                    # 清除会话
                    del self.user_sessions[from_wxid]
                    # 执行抓取流程
                    self._start_crawl_shops(instance_wxid, from_wxid, mt_id)
                    
                else:
                    # 无效选择，提示重新选择
                    self._safe_send_text(instance_wxid, from_wxid, "❌ 无效选择，请发送 \"1\" 或 \"2\"\n\n如需帮助请发送\"教程\"查看使用教程")
            
            # 处理查返利 - 等待token URL
            elif session_type == "rebate_wait_token":
                self._handle_rebate_token_input(instance_wxid, from_wxid, choice)
            
            # 处理查返利 - 等待订单号
            elif session_type == "rebate_wait_order":
                self._handle_rebate_order_input(instance_wxid, from_wxid, choice)
            
            # 处理联盟榜单 - 等待榜单类型选择
            elif session_type == "alliance_wait_list_type":
                self._handle_alliance_list_type_choice(instance_wxid, from_wxid, choice)
            
            # 处理联盟榜单 - 等待业务线选择
            elif session_type == "alliance_wait_biz_line":
                self._handle_alliance_biz_line_choice(instance_wxid, from_wxid, choice)
            
            # 处理联盟榜单 - 等待城市输入
            elif session_type == "alliance_wait_city":
                self._handle_alliance_city_input(instance_wxid, from_wxid, choice)
            
            # 处理企业返利 - 等待SK输入
            elif session_type == "enterprise_rebate_wait_sk":
                self._handle_enterprise_sk_input(instance_wxid, from_wxid, choice)
            
            # 处理企业返利 - 等待订单号
            elif session_type == "enterprise_rebate_wait_order":
                self._handle_enterprise_order_input(instance_wxid, from_wxid, choice)
                
        except Exception as e:
            self.add_log("ERROR", f"处理用户选择时出错: {e}")
            # 清除会话
            if from_wxid in self.user_sessions:
                del self.user_sessions[from_wxid]
    
    def _start_crawl_shops(self, instance_wxid, from_wxid, mt_id):
        """开始抓取门店流程"""
        try:
            # 检查openpyxl是否可用
            if not OPENPYXL_AVAILABLE:
                self.add_log("ERROR", "缺少openpyxl库，无法生成Excel文件")
                self._safe_send_text(instance_wxid, from_wxid, "系统错误：缺少必要的库，请联系管理员")
                return
            
            # 检查用户权限（再次确认）
            has_permission, error_msg, user = self._check_user_permission(from_wxid)
            if not has_permission:
                self._safe_send_text(instance_wxid, from_wxid, error_msg)
                return
            
            # 检查会员状态
            is_vip, vip_type, expire_time = self._check_vip_status(from_wxid)
            
            # 检查本地是否已有该美团ID的文件（当天内）
            cached_file = self._check_local_file(mt_id)
            if cached_file:
                self.add_log("INFO", f"使用本地缓存文件: {cached_file}")
                self._safe_send_text(instance_wxid, from_wxid, "请稍后，正在抓取中...")
                
                # 直接发送文件（添加延迟）
                time.sleep(0.5)
                result = self.send_file(instance_wxid, from_wxid, cached_file)
                
                if result.get("success"):
                    self.add_log("INFO", f"缓存文件发送成功: {cached_file}")
                    
                    # 从缓存文件中读取门店数据并发送统计信息
                    shops = self._read_shops_from_excel(cached_file)
                    if shops:
                        self._send_shop_statistics(instance_wxid, from_wxid, mt_id, shops)
                    
                    # 从缓存发送不计次数、不扣费
                    self.add_log("INFO", f"用户 {from_wxid} 使用缓存文件，不计入次数和扣费")
                else:
                    error = result.get("error", "未知错误")
                    self.add_log("ERROR", f"缓存文件发送失败: {error}")
                    self._safe_send_text(instance_wxid, from_wxid, f"文件发送失败: {error}")
                return
            
            # 没有缓存文件，继续抓取流程
            # 检查非会员用户的使用次数和余额（抓取成功后再扣费）
            cost = 0.0  # 抓取门店费用
            if not is_vip:
                # 非会员用户，检查今日使用次数（1次免费，超过1次每次扣0.5元）
                daily_usage = self._get_daily_usage(from_wxid)
                if daily_usage >= 1:
                    # 超过1次，需要扣费
                    cost = 0.5
                    # 检查余额是否足够
                    if user["balance"] < cost:
                        self._safe_send_text(instance_wxid, from_wxid, f"余额不足！当前余额: {user['balance']:.2f}元，需要: {cost:.2f}元")
                        return
                    self.add_log("INFO", f"用户 {from_wxid} 今日已使用{daily_usage}次（超过1次免费额度），本次抓取成功后需扣费 {cost}元")
                else:
                    self.add_log("INFO", f"用户 {from_wxid} 今日已使用{daily_usage}次，在免费额度内")
            else:
                # 会员用户不受次数限制，不扣费
                self.add_log("INFO", f"用户 {from_wxid} 是{vip_type}会员，不受次数限制，不扣费")
            
            # 回复用户
            self._safe_send_text(instance_wxid, from_wxid, "请稍后，正在抓取中...")
            
            # 记录任务信息用于后续回复
            task_info = {
                "instance_wxid": instance_wxid,
                "from_wxid": from_wxid,
                "mt_id": mt_id,
                "start_time": datetime.now(),
                "cost": cost,  # 传递扣费信息
                "is_vip": is_vip  # 传递会员状态
            }
            
            # 启动子线程处理
            thread = threading.Thread(
                target=self._process_meituan_shop,
                args=(task_info,),
                daemon=True
            )
            thread.start()
            
        except Exception as e:
            self.add_log("ERROR", f"启动抓取流程时出错: {e}")
            try:
                self._safe_send_text(instance_wxid, from_wxid, f"处理出错: {str(e)}")
            except:
                pass
    
    def _process_meituan_shop(self, task_info):
        """在子线程中处理美团门店抓取"""
        instance_wxid = task_info["instance_wxid"]
        from_wxid = task_info["from_wxid"]
        mt_id = task_info["mt_id"]
        cost = task_info.get("cost", 0.0)  # 获取扣费金额
        is_vip = task_info.get("is_vip", False)  # 获取会员状态
        
        try:
            self.add_log("INFO", f"开始抓取美团ID {mt_id} 的门店列表")
            
            # 调用获取门店列表函数
            shops = get_sku_shop(
                limit=self.mt_limit,
                token=self.mt_token,
                sku=int(mt_id),
                offset=0,
                _safe_send_text= self._safe_send_text,
                instance_wxid = instance_wxid
            )
            
            if not shops:
                self.add_log("WARNING", f"美团ID {mt_id} 未抓取到任何门店")
                self._safe_send_text(instance_wxid, from_wxid, "抱歉，未能抓取到门店信息，请检查美团ID是否正确")
                return
        
        except MeituanRiskControlException as e:
            # 遇到风控，不保存文件，不发送给客户
            self.add_log("ERROR", f"抓取门店时遇到风控: {str(e)}")
            
            # 发送提示消息给客户
            self._safe_send_text(instance_wxid, from_wxid, "抱歉，遇见风控，已通知管理员处理，请稍后重试。")
            
            # 通知管理员（发送 generalPageUrl）
            admin_wxid = "wxid_3intiqznkov222"
            if e.general_page_url:
                notification_msg = f"🚨 风控通知\n美团ID: {mt_id}\n用户: {from_wxid}\n风控验证URL:\n{e.general_page_url}"
                self._safe_send_text(instance_wxid, admin_wxid, notification_msg)
                self.add_log("INFO", f"已通知管理员风控情况，URL: {e.general_page_url}")
            else:
                notification_msg = f"🚨 风控通知\n美团ID: {mt_id}\n用户: {from_wxid}\n遇到风控但未获取到验证URL"
                self._safe_send_text(instance_wxid, admin_wxid, notification_msg)
                self.add_log("WARNING", "风控异常但未获取到 generalPageUrl")
            
            return
        
        except Exception as e:
            self.add_log("ERROR", f"抓取门店时出错: {e}")
            try:
                self._safe_send_text(instance_wxid, from_wxid, f"抓取失败: {str(e)}")
            except:
                pass
            return
        
        # 正常处理抓取到的门店数据
        self.add_log("INFO", f"成功抓取到 {len(shops)} 家门店")
        
        # 删除旧的缓存文件
        try:
            if os.path.exists(self.excel_output_dir):
                files = os.listdir(self.excel_output_dir)
                for filename in files:
                    if filename.endswith('.xlsx') and f"-{mt_id}-" in filename:
                        old_file_path = os.path.join(self.excel_output_dir, filename)
                        os.remove(old_file_path)
                        self.add_log("INFO", f"已删除旧缓存文件: {old_file_path}")
        except Exception as e:
            self.add_log("WARNING", f"删除旧缓存文件时出错: {e}")
        
        # 生成Excel文件
        file_path = self._save_shops_to_excel(mt_id, shops)
        
        if not file_path:
            self.add_log("ERROR", "生成Excel文件失败")
            self._safe_send_text(instance_wxid, from_wxid, "抱歉，生成文件失败，请稍后重试")
            return
        
        self.add_log("INFO", f"Excel文件已生成: {file_path}")
        
        # 发送文件给用户（添加延迟）
        time.sleep(0.5)
        
        result = self.send_file(instance_wxid, from_wxid, file_path)
        
        if result.get("success"):
            self.add_log("INFO", f"文件发送成功: {file_path}")
            
            # 发送门店统计信息
            self._send_shop_statistics(instance_wxid, from_wxid, mt_id, shops)
            
            # 抓取成功，执行扣费和增加使用次数
            if not is_vip:
                # 非会员用户，先扣费再增加使用次数
                if cost > 0:
                    if self._deduct_balance(from_wxid, cost):
                        self.add_log("INFO", f"已扣除用户 {from_wxid} 余额 {cost:.2f}元")
                    else:
                        self.add_log("ERROR", f"扣除用户 {from_wxid} 余额失败")
                        # 扣费失败也要记录，避免反复扣费失败
                
                # 增加今日使用次数
                self._increment_daily_usage(from_wxid)
                current_usage = self._get_daily_usage(from_wxid)
                self.add_log("INFO", f"用户 {from_wxid} 今日使用次数: {current_usage}次（1次免费，超过1次每次扣0.5元）")
            else:
                self.add_log("INFO", f"用户 {from_wxid} 是会员，不计入次数不扣费")
            
            # 根据配置决定是否删除文件
            if self.delete_after_send:
                try:
                    os.remove(file_path)
                    self.add_log("INFO", f"已删除临时文件: {file_path}")
                except Exception as e:
                    self.add_log("WARNING", f"删除文件失败: {e}")
            else:
                self.add_log("INFO", f"文件已保存到: {file_path}")
        else:
            error = result.get("error", "未知错误")
            self.add_log("ERROR", f"文件发送失败: {error}")
            self._safe_send_text(instance_wxid, from_wxid, f"文件发送失败: {error}")
    
    def _save_statistics_to_txt(self, mt_id, statistics_msg, brand_name):
        """
        将统计信息保存为txt文件
        
        Args:
            mt_id: 美团ID
            statistics_msg: 统计信息文本
            brand_name: 品牌名称
            
        Returns:
            str: 文件路径，失败返回None
        """
        try:
            # 生成文件名：品牌名-统计信息.txt
            filename = f"{brand_name}-统计信息.txt"
            
            # 保存到专门的Excel输出目录
            file_path = os.path.join(self.excel_output_dir, filename)
            
            # 确保目录存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # 写入文件
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(statistics_msg)
            
            self.add_log("INFO", f"统计信息已保存到文件: {file_path}")
            return file_path
            
        except Exception as e:
            self.add_log("ERROR", f"保存统计信息到文件时出错: {e}")
            import traceback
            self.add_log("ERROR", f"详细错误: {traceback.format_exc()}")
            return None
    
    def _send_shop_statistics(self, instance_wxid, from_wxid, mt_id, shops):
        """
        发送门店统计信息（保存为txt文件发送）
        
        Args:
            instance_wxid: 微信实例ID
            from_wxid: 发送者微信ID
            mt_id: 美团ID
            shops: 门店列表
        """
        try:
            if not shops or len(shops) == 0:
                return
            
            # 生成统计信息
            statistics_msg = self._generate_shop_statistics(mt_id, shops)
            if not statistics_msg:
                return
            
            # 提取品牌名
            brand_name = "未知品牌"
            if shops and len(shops) > 0:
                first_shop_name = shops[0].get("name", "")
                brand_name = self._extract_brand_name(first_shop_name)
            
            # 保存为txt文件发送
            shop_count = len(shops)
            self.add_log("INFO", f"门店数量({shop_count})，保存为txt文件发送")
            txt_file_path = self._save_statistics_to_txt(mt_id, statistics_msg, brand_name)
            
            if txt_file_path:
                time.sleep(0.5)
                result = self.send_file(instance_wxid, from_wxid, txt_file_path)
                
                if result.get("success"):
                    self.add_log("INFO", f"统计信息文件发送成功: {txt_file_path}")
                    
                    # 根据配置决定是否删除文件
                    if self.delete_after_send:
                        try:
                            os.remove(txt_file_path)
                            self.add_log("INFO", f"已删除统计信息文件: {txt_file_path}")
                        except Exception as e:
                            self.add_log("WARNING", f"删除文件失败: {e}")
                else:
                    error = result.get("error", "未知错误")
                    self.add_log("ERROR", f"统计信息文件发送失败: {error}")
            else:
                self.add_log("ERROR", "生成统计信息文件失败")
                
        except Exception as e:
            self.add_log("ERROR", f"发送门店统计信息时出错: {e}")
            import traceback
            self.add_log("ERROR", f"详细错误: {traceback.format_exc()}")
    
    def _read_shops_from_excel(self, file_path):
        """
        从Excel文件中读取门店数据
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            list: 门店列表 [{"name": "门店名称", "cityName": "城市"}]
        """
        try:
            if not OPENPYXL_AVAILABLE:
                self.add_log("ERROR", "缺少openpyxl库，无法读取Excel文件")
                return None

            from openpyxl import load_workbook
            
            # 加载工作簿
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            
            shops = []
            # 跳过第一行（表头）
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2:
                    shop_name = row[0]
                    city_name = row[1]
                    if shop_name and city_name:
                        shops.append({
                            "name": str(shop_name),
                            "cityName": str(city_name)
                        })
            
            wb.close()
            self.add_log("INFO", f"从Excel文件读取了{len(shops)}家门店")
            return shops
            
        except Exception as e:
            self.add_log("ERROR", f"读取Excel文件时出错: {e}")
            import traceback
            self.add_log("ERROR", f"详细错误: {traceback.format_exc()}")
            return None
    
    def _generate_shop_statistics(self, mt_id, shops):
        """
        生成门店统计信息
        
        Args:
            mt_id: 美团ID
            shops: 门店列表 [{"name": "门店名称", "cityName": "城市"}]
            
        Returns:
            str: 统计信息文本
        """
        try:
            if not shops or len(shops) == 0:
                return None
            
            # 提取品牌名
            brand_name = "未知品牌"
            if shops and len(shops) > 0:
                first_shop_name = shops[0].get("name", "")
                brand_name = self._extract_brand_name(first_shop_name)
            
            # 统计城市和门店
            city_shops = {}  # {城市: set(门店列表)} 使用set自动去重
            for shop in shops:
                city = shop.get("cityName", "未知城市")
                shop_name = shop.get("name", "")
                
                # 简化门店名称（去掉品牌前缀）
                simplified_name = shop_name
                if brand_name and brand_name in shop_name:
                    # 尝试提取括号内的内容
                    import re
                    match = re.search(r'[（(](.+?)[）)]', shop_name)
                    if match:
                        simplified_name = match.group(1)
                    else:
                        # 如果没有括号，直接去掉品牌名
                        simplified_name = shop_name.replace(brand_name, "").strip()
                
                if city not in city_shops:
                    city_shops[city] = set()  # 使用set去重
                city_shops[city].add(simplified_name)  # 使用add而不是append
            
            # 计算统计数据
            city_count = len(city_shops)
            # 使用原始门店总数（不去重）
            shop_count = len(shops)
            
            # 构建统计信息
            lines = []
            lines.append(f"{brand_name}|{mt_id}")
            lines.append(f"{city_count}座城市    {shop_count}家门店可用")
            lines.append("=" * 16)
            lines.append("可用城市：")
            
            # 城市列表（用顿号分隔）
            cities = list(city_shops.keys())
            lines.append("、".join(cities))
            lines.append("=" * 16)
            lines.append("可用门店：")
            
            # 按城市分组的门店列表
            for city in cities:
                shop_names = list(city_shops[city])  # 将set转为list
                # 用顿号分隔门店名
                shops_text = "、".join(shop_names)
                lines.append(f"【{city}】：{shops_text}")
            
            return "\n".join(lines)
            
        except Exception as e:
            self.add_log("ERROR", f"生成门店统计信息时出错: {e}")
            import traceback
            self.add_log("ERROR", f"详细错误: {traceback.format_exc()}")
            return None
    
    def _check_local_file(self, mt_id, max_days=3):
        """
        检查本地是否存在指定美团ID的文件（3天内）
        
        Args:
            mt_id: 美团ID
            max_days: 缓存文件有效天数，默认3天
            
        Returns:
            str: 文件路径，如果不存在或已过期返回None
        """
        try:
            if not os.path.exists(self.excel_output_dir):
                return None
            
            # 获取N天前0点的时间戳
            days_ago_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=max_days-1)
            days_ago_start_timestamp = days_ago_start.timestamp()
            
            self.add_log("INFO", f"检查本地文件，{max_days}天前0点时间戳: {days_ago_start_timestamp}")
            
            # 获取所有文件
            files = os.listdir(self.excel_output_dir)
            
            # 过滤出包含该美团ID的xlsx文件
            # 文件名格式：品牌名-美团ID-时间戳.xlsx
            matching_files = []
            for filename in files:
                if filename.endswith('.xlsx') and f"-{mt_id}-" in filename:
                    file_path = os.path.join(self.excel_output_dir, filename)
                    
                    # 从文件名中提取时间戳
                    try:
                        # 文件名格式：品牌名-美团ID-时间戳.xlsx
                        parts = filename.rsplit('-', 1)  # 从右侧分割一次
                        if len(parts) == 2:
                            timestamp_str = parts[1].replace('.xlsx', '')
                            file_timestamp = int(timestamp_str)
                            
                            # 判断文件是否在有效期内（N天前0点之后创建）
                            if file_timestamp >= days_ago_start_timestamp:
                                matching_files.append((file_path, file_timestamp))
                                self.add_log("INFO", f"文件 {filename} 在{max_days}天有效期内，文件时间戳: {file_timestamp}")
                            else:
                                self.add_log("INFO", f"文件 {filename} 已过期（超过{max_days}天），跳过")
                    except (ValueError, IndexError) as e:
                        # 如果无法解析时间戳，使用文件修改时间作为fallback
                        self.add_log("WARNING", f"无法从文件名中解析时间戳: {filename}, 使用文件修改时间")
                        file_mtime = os.path.getmtime(file_path)
                        
                        # 判断文件修改时间是否在有效期内
                        if file_mtime >= days_ago_start_timestamp:
                            matching_files.append((file_path, file_mtime))
            
            # 如果有多个文件，返回最新的
            if matching_files:
                matching_files.sort(key=lambda x: x[1], reverse=True)
                latest_file = matching_files[0][0]
                self.add_log("INFO", f"找到本地缓存文件: {latest_file}")
                return latest_file
            
            return None
            
        except Exception as e:
            self.add_log("ERROR", f"检查本地文件时出错: {e}")
            return None
    
    def _extract_brand_name(self, shop_name):
        """
        从门店名称中提取品牌名（括号前的部分）
        
        Args:
            shop_name: 门店名称，如 "牧牛煌无限量点餐烤肉（吕梁兴县店）"
            
        Returns:
            str: 品牌名，如 "牧牛煌无限量点餐烤肉"
        """
        if not shop_name:
            return "未知品牌"
        
        # 查找各种括号
        brackets = ['(', '（', '[', '【']
        for bracket in brackets:
            if bracket in shop_name:
                brand_name = shop_name.split(bracket)[0].strip()
                if brand_name:
                    return brand_name
        
        # 如果没有括号，返回原名称
        return shop_name.strip()
    
    def _save_shops_to_excel(self, mt_id, shops):
        """
        将门店列表保存为Excel文件
        
        Args:
            mt_id: 美团ID
            shops: 门店列表 [{"name": "门店名称", "cityName": "城市"}]
            
        Returns:
            str: 文件路径，失败返回None
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "门店列表"
            
            # 写入表头
            ws.append(["门店名称", "城市"])
            
            # 写入数据
            for shop in shops:
                ws.append([shop.get("name", ""), shop.get("cityName", "")])
            
            # 调整列宽
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 20
            
            # 从第一个门店名称中提取品牌名
            brand_name = "未知品牌"
            if shops and len(shops) > 0:
                first_shop_name = shops[0].get("name", "")
                brand_name = self._extract_brand_name(first_shop_name)
            
            # 生成时间戳（Unix时间戳）
            timestamp = str(int(time.time()))
            
            # 生成文件名：品牌名-美团ID-时间戳.xlsx
            filename = f"{brand_name}-{mt_id}-{timestamp}.xlsx"
            
            # 保存到专门的Excel输出目录
            file_path = os.path.join(self.excel_output_dir, filename)
            
            # 确保目录存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # 保存文件
            wb.save(file_path)
            
            self.add_log("INFO", f"Excel文件已保存到本地: {file_path}")
            
            return file_path
        
        except Exception as e:
            self.add_log("ERROR", f"保存Excel文件时出错: {e}")
            import traceback
            self.add_log("ERROR", f"详细错误: {traceback.format_exc()}")
            return None
    
    def _save_rebate_results_to_excel(self, query_results):
        """
        将返利查询结果保存为Excel文件
        
        Args:
            query_results: 查询结果列表 [{"order_id": "xxx", "status": "success", "detail": {...}}]
            
        Returns:
            str: 文件路径，失败返回None
        """
        try:
            # 检查openpyxl是否可用
            if not OPENPYXL_AVAILABLE:
                self.add_log("ERROR", "缺少openpyxl库，无法生成Excel文件")
                return None
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "返利查询结果"
            
            # 写入表头
            ws.append(["订单号", "状态", "推广单号", "佣金", "卡券状态", "下单时间", "下单城市", "核销城市"])
            
            # 写入数据
            for item in query_results:
                order_id = item.get("order_id", "")
                status = item.get("status", "")
                status_text = "走返利成功" if status == "success" else "走返利失败"
                
                detail = item.get("detail")
                if detail:
                    # 转换卡券状态
                    order_status = detail.get("orderStatus", "")
                    if order_status == 3:
                        status_desc = "已退款"
                    elif order_status == 4:
                        status_desc = "未核销"
                    elif order_status == 5:
                        status_desc = "已核销"
                    else:
                        status_desc = str(order_status) if order_status else "未知"
                    
                    # 转换下单时间
                    order_pay_time = detail.get("orderPayTime", "")
                    if order_pay_time:
                        try:
                            if isinstance(order_pay_time, int):
                                time_str = datetime.fromtimestamp(order_pay_time).strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                time_str = str(order_pay_time)
                        except:
                            time_str = str(order_pay_time)
                    else:
                        time_str = ""
                    
                    ws.append([
                        order_id,
                        status_text,
                        detail.get("orderViewId", ""),
                        detail.get("commissionFee", ""),
                        status_desc,
                        time_str,
                        detail.get("cityName", ""),
                        detail.get("consumeCityName", "") if detail.get("consumeCityName") else ""
                    ])
                else:
                    # 没有详细信息的订单（走返利失败）
                    ws.append([order_id, status_text, "", "", "", "", "", ""])
            
            # 调整列宽
            ws.column_dimensions['A'].width = 20  # 订单号
            ws.column_dimensions['B'].width = 15  # 状态
            ws.column_dimensions['C'].width = 30  # 推广单号
            ws.column_dimensions['D'].width = 10  # 佣金
            ws.column_dimensions['E'].width = 12  # 卡券状态
            ws.column_dimensions['F'].width = 20  # 下单时间
            ws.column_dimensions['G'].width = 15  # 下单城市
            ws.column_dimensions['H'].width = 15  # 核销城市
            
            # 生成时间戳（Unix时间戳）
            timestamp = str(int(time.time()))
            
            # 生成文件名：返利查询结果.xlsx
            filename = f"返利查询结果.xlsx"
            
            # 保存到专门的Excel输出目录
            file_path = os.path.join(self.excel_output_dir, filename)
            
            # 确保目录存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # 保存文件
            wb.save(file_path)
            
            self.add_log("INFO", f"返利查询Excel文件已保存到本地: {file_path}")
            
            return file_path
        
        except Exception as e:
            self.add_log("ERROR", f"保存返利查询Excel文件时出错: {e}")
            import traceback
            self.add_log("ERROR", f"详细错误: {traceback.format_exc()}")
            return None

